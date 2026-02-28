from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    send_file,
    send_from_directory,
    flash,
    session,
    jsonify,
    g,
    current_app,
    Response,
    stream_with_context,
    has_request_context,
)
from flask_migrate import Migrate, upgrade
import logging
from logging.handlers import RotatingFileHandler
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from flask_wtf import CSRFProtect
from models import (
    db,
    migrate,
    Client,
    Product,
    Quotation,
    QuotationItem,
    Order,
    OrderItem,
    Invoice,
    InvoiceItem,
    Payment,
    InventoryMovement,
    Warehouse,
    ProductStock,
    ProductPriceLog,
    CompanyInfo,
    User,
    AccountRequest,
    ExportLog,
    NcfLog,
    Notification,
    SystemAnnouncement,
    ErrorReport,
    AuditLog,
    AppSetting,
    RNCRegistry,
    dom_now,
)
from io import BytesIO, StringIO
from urllib.parse import quote_plus
import csv
try:
    from openpyxl import Workbook
except ModuleNotFoundError:  # pragma: no cover
    Workbook = None
from datetime import datetime, timedelta
from pathlib import Path
from sqlalchemy import and_, extract, func, inspect, or_
from sqlalchemy.exc import NoSuchTableError
from sqlalchemy.orm import load_only, joinedload
from sqlalchemy.engine import make_url
from werkzeug.utils import secure_filename
from werkzeug.exceptions import HTTPException
from werkzeug.security import generate_password_hash
import os
import re
import json
import uuid
import importlib.util
import hashlib
from ai import recommend_products
from weasy_pdf import generate_pdf, generate_pdf_bytes
from account_pdf import generate_account_statement_pdf, generate_account_statement_pdf_bytes
from functools import wraps
from auth import auth_bp, generate_reset_token
from forms import AccountRequestForm
from config import DevelopmentConfig, TestingConfig, ProductionConfig, validate_runtime_config
try:
    from dotenv import load_dotenv
except ModuleNotFoundError:  # pragma: no cover
    def load_dotenv():
        pass
try:
    from rq import Queue
    from redis import Redis
except Exception:  # pragma: no cover
    Queue = None
    Redis = None
import threading
from queue import Queue as ThreadQueue, Empty
import time
import random

load_dotenv()

# Load RNC data for company name lookup
RNC_DATA = {}
DATA_PATH = os.path.join(os.path.dirname(__file__), 'data', 'DGII_RNC.TXT')
def _parse_rnc_line(raw_line: str) -> tuple[str, str] | tuple[None, None]:
    line = (raw_line or '').strip()
    if not line:
        return None, None
    parts = [p.strip() for p in line.split('|') if p is not None]
    if len(parts) >= 2:
        rnc = re.sub(r'\D', '', parts[0])
        name = parts[1].strip()
        return (rnc, name) if rnc and name else (None, None)
    # Fallback para archivos separados por tab/espacios
    bits = line.split()
    if len(bits) >= 2:
        rnc = re.sub(r'\D', '', bits[0])
        name = ' '.join(bits[1:]).strip()
        return (rnc, name) if rnc and name else (None, None)
    return None, None


if os.path.exists(DATA_PATH):
    with open(DATA_PATH, encoding='utf-8') as f:
        for row in f:
            rnc, name = _parse_rnc_line(row)
            if rnc and name:
                RNC_DATA[rnc] = name

app = Flask(__name__)
APP_ENV = os.getenv('APP_ENV', 'development').strip().lower()


APP_VERSION = os.getenv('APP_VERSION', '2.0.5').strip() or '2.0.5'
APP_VERSION_HIGHLIGHTS = [
    'Conexión MySQL/cPanel simplificada con soporte DB_* y DATABASE_URL.',
    'Reportar Problema, anuncios del sistema y panel administrativo ampliado.',
    'Renderizado PDF con fpdf2 para descargas más estables en hosting compartido.',
    'Refuerzo de seguridad: validación runtime, bloqueo de rutas sensibles y mejoras de auditoría.',
]

def _normalized_database_url(url: str | None) -> str | None:
    """Normalize DB URL so cPanel/MySQL strings work with SQLAlchemy.

    cPanel users often set `mysql://...`; SQLAlchemy expects an explicit driver
    such as `mysql+pymysql://...`.
    """
    if not url:
        return url
    if url.startswith('mysql://'):
        return url.replace('mysql://', 'mysql+pymysql://', 1)
    return url


def _database_url_from_parts() -> str | None:
    """Build DATABASE_URL from simple DB_* env vars for easier cPanel setup."""
    db_name = (os.getenv('DB_NAME') or '').strip()
    db_user = (os.getenv('DB_USER') or '').strip()
    db_password = os.getenv('DB_PASSWORD')
    if not db_name or not db_user or db_password is None:
        return None

    db_driver = (os.getenv('DB_DRIVER', 'mysql+pymysql') or 'mysql+pymysql').strip()
    db_host = (os.getenv('DB_HOST', 'localhost') or 'localhost').strip()
    db_port = (os.getenv('DB_PORT') or '').strip()

    host_part = db_host
    if db_port:
        host_part = f"{db_host}:{db_port}"

    quoted_user = quote_plus(db_user)
    quoted_password = quote_plus(db_password)
    quoted_db_name = quote_plus(db_name)
    return f"{db_driver}://{quoted_user}:{quoted_password}@{host_part}/{quoted_db_name}?charset=utf8mb4"


def _resolve_database_url() -> tuple[str | None, str]:
    database_url = os.getenv('DATABASE_URL')
    source = 'DATABASE_URL'
    if not database_url:
        database_url = _database_url_from_parts()
        source = 'DB_* variables' if database_url else 'default config'
    return _normalized_database_url(database_url), source


database_url, database_url_source = _resolve_database_url()
if database_url:
    os.environ['DATABASE_URL'] = database_url


def _apply_database_uri_override(config: dict, resolved_url: str | None):
    """Ensure runtime config uses resolved DB URL even after class import-time defaults."""
    if resolved_url:
        config['SQLALCHEMY_DATABASE_URI'] = resolved_url


config_map = {
    'development': DevelopmentConfig,
    'testing': TestingConfig,
    'production': ProductionConfig,
}
app.config.from_object(config_map.get(APP_ENV, DevelopmentConfig))
_apply_database_uri_override(app.config, database_url)
app.config['APP_ENV'] = APP_ENV
validate_runtime_config(app.config)

SENSITIVE_PATHS = {
    'app.py',
    'auth.py',
    'config.py',
    'models.py',
    'requirements.txt',
    '.env',
    '.env.example',
    'database.sqlite',
    'passenger_wsgi.py',
}
SENSITIVE_EXTENSIONS = (
    '.py', '.pyc', '.sqlite', '.db', '.env', '.ini', '.pem', '.key', '.log',
)


@app.before_request
def block_sensitive_paths():
    requested = request.path.lstrip('/').lower()
    if not requested:
        return None
    if requested in SENSITIVE_PATHS or requested.endswith(SENSITIVE_EXTENSIONS):
        return ('Not Found', 404)
    return None


@app.before_request
def attach_request_context_log():
    g.request_id = uuid.uuid4().hex[:12]
    g.request_started_at = time.time()
    app.logger.info(
        'REQ_START id=%s method=%s path=%s ip=%s ua=%s',
        g.request_id,
        request.method,
        request.full_path,
        request.remote_addr,
        request.user_agent.string[:200] if request.user_agent else '',
    )


@app.after_request
def log_response_result(response):
    rid = getattr(g, 'request_id', '-')
    started = getattr(g, 'request_started_at', None)
    duration_ms = int((time.time() - started) * 1000) if started else -1
    app.logger.info(
        'REQ_END id=%s status=%s duration_ms=%s',
        rid,
        response.status_code,
        duration_ms,
    )
    return response


@app.errorhandler(Exception)
def handle_unexpected_error(exc):
    if isinstance(exc, HTTPException):
        return exc
    rid = getattr(g, 'request_id', '-')
    app.logger.exception('REQ_FAIL id=%s method=%s path=%s err=%s', rid, request.method, request.path, exc)
    return render_template('error.html', request_id=rid), 500


if not os.path.exists('logs'):
    os.makedirs('logs')
file_handler = RotatingFileHandler('logs/app.log', maxBytes=10240, backupCount=10)
file_handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'))
file_handler.setLevel(logging.INFO)
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)
app.logger.info('Tiendix startup')
if database_url:
    app.logger.info('Database configuration loaded from %s', database_url_source)
    app.logger.info('Database URI override applied from resolved environment variables')
else:
    app.logger.info('Database configuration loaded from default config')

MAIL_SERVER = os.getenv('MAIL_SERVER')
MAIL_PORT = int(os.getenv('MAIL_PORT', 587))
MAIL_USERNAME = os.getenv('MAIL_USERNAME')
MAIL_PASSWORD = os.getenv('MAIL_PASSWORD')
MAIL_DEFAULT_SENDER = os.getenv('MAIL_DEFAULT_SENDER', MAIL_USERNAME)
MAIL_MAX_RETRIES = int(os.getenv('MAIL_MAX_RETRIES', 3))
MAIL_RETRY_DELAY_SEC = float(os.getenv('MAIL_RETRY_DELAY_SEC', 1))

EMAIL_METRICS = {
    'queued': 0,
    'sent': 0,
    'failed': 0,
    'retries': 0,
    'skipped': 0,
}
_email_queue = ThreadQueue()


def _deliver_email(to, subject, html, attachments=None):
    if not MAIL_SERVER or not MAIL_DEFAULT_SENDER:
        EMAIL_METRICS['skipped'] += 1
        app.logger.warning('Email settings missing; skipping send to %s', to)
        return

    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From'] = MAIL_DEFAULT_SENDER
    msg['To'] = to
    msg.attach(MIMEText(html, 'html'))
    for attachment in attachments or []:
        filename, data = attachment
        part = MIMEApplication(data, Name=filename)
        part['Content-Disposition'] = f'attachment; filename="{filename}"'
        msg.attach(part)

    for attempt in range(1, MAIL_MAX_RETRIES + 1):
        try:
            with smtplib.SMTP(MAIL_SERVER, MAIL_PORT) as s:
                if MAIL_USERNAME and MAIL_PASSWORD:
                    s.starttls()
                    s.login(MAIL_USERNAME, MAIL_PASSWORD)
                s.sendmail(MAIL_DEFAULT_SENDER, [to], msg.as_string())
            EMAIL_METRICS['sent'] += 1
            return
        except Exception as e:  # pragma: no cover
            if attempt < MAIL_MAX_RETRIES:
                EMAIL_METRICS['retries'] += 1
                app.logger.warning('Email send failed (attempt %s/%s) to %s: %s', attempt, MAIL_MAX_RETRIES, to, e)
                time.sleep(MAIL_RETRY_DELAY_SEC)
                continue
            EMAIL_METRICS['failed'] += 1
            app.logger.error('Email send failed after %s attempts to %s: %s', MAIL_MAX_RETRIES, to, e)


def _email_worker():  # pragma: no cover - background helper
    while True:
        try:
            payload = _email_queue.get(timeout=1)
        except Empty:
            continue
        if payload is None:
            break
        _deliver_email(*payload)
        _email_queue.task_done()


_email_worker_thread = threading.Thread(target=_email_worker, daemon=True)
_email_worker_thread.start()


def send_email(to, subject, html, attachments=None, asynchronous=True):
    if asynchronous:
        EMAIL_METRICS['queued'] += 1
        _email_queue.put((to, subject, html, attachments))
        return
    _deliver_email(to, subject, html, attachments)


def _fmt_money(value):
    return f"RD${value:,.2f}"


app.jinja_env.filters['money'] = _fmt_money



def _module_available(module_name: str) -> bool:
    try:
        return importlib.util.find_spec(module_name) is not None
    except (ModuleNotFoundError, ValueError):
        return False


def _ensure_mysql_driver_available(config: dict):
    """Ensure selected MySQL driver exists; fallback to available alternatives."""
    uri = config.get('SQLALCHEMY_DATABASE_URI') or ''
    if not isinstance(uri, str):
        return
    if not uri.startswith('mysql+pymysql://'):
        return

    if _module_available('pymysql'):
        return

    if _module_available('MySQLdb'):
        config['SQLALCHEMY_DATABASE_URI'] = uri.replace('mysql+pymysql://', 'mysql+mysqldb://', 1)
        app.logger.warning('PyMySQL not installed; falling back to mysql+mysqldb driver')
        return

    app.logger.error(
        'MySQL driver not found (pymysql/mysqldb). '
        'Falling back to sqlite temporarily. Install PyMySQL in cPanel virtualenv and restart.'
    )
    config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.sqlite'



def _maybe_fix_cpanel_access_denied(config: dict):
    """Try DB_NAME=DB_USER fallback for common cPanel grant mismatch (1044)."""
    uri = config.get('SQLALCHEMY_DATABASE_URI') or ''
    if not isinstance(uri, str) or not uri.startswith('mysql+'):
        return
    if not _module_available('pymysql'):
        return

    db_user_env = (os.getenv('DB_USER') or '').strip()
    if not db_user_env:
        return

    try:
        parsed = make_url(uri)
    except Exception:
        return

    current_db = (parsed.database or '').strip()
    if not current_db or current_db == db_user_env:
        return

    import pymysql

    connect_kwargs = {
        'host': parsed.host or 'localhost',
        'port': int(parsed.port or 3306),
        'user': parsed.username,
        'password': parsed.password,
        'database': current_db,
        'connect_timeout': 4,
    }

    try:
        conn = pymysql.connect(**connect_kwargs)
        conn.close()
        return
    except Exception as exc:
        err_text = str(exc)
        # 1044 is common when user exists but has no grants on selected DB.
        if '1044' not in err_text and 'Access denied for user' not in err_text:
            return

    try:
        fallback_kwargs = dict(connect_kwargs)
        fallback_kwargs['database'] = db_user_env
        conn = pymysql.connect(**fallback_kwargs)
        conn.close()
        fallback_uri = parsed.set(database=db_user_env).render_as_string(hide_password=False)
        config['SQLALCHEMY_DATABASE_URI'] = fallback_uri
        app.logger.warning(
            'DB access denied for %s; switched DB_NAME to DB_USER fallback: %s',
            current_db,
            db_user_env,
        )
    except Exception:
        # Keep original URI if fallback also fails.
        return

_ensure_mysql_driver_available(app.config)
_maybe_fix_cpanel_access_denied(app.config)
db.init_app(app)
migrate.init_app(app, db)
csrf = CSRFProtect(app)
app.register_blueprint(auth_bp)

# The database schema is managed via Flask-Migrate.  Tables should be
# created with ``flask db upgrade`` instead of ``db.create_all`` to avoid
# diverging from migrations.

if Queue and Redis:
    redis_conn = Redis.from_url(os.getenv('REDIS_URL', 'redis://localhost:6379'))
    export_queue = Queue('exports', connection=redis_conn)
else:  # pragma: no cover
    export_queue = None


def enqueue_export(fn, *args):
    """Enqueue an export job using RQ if available or fallback to threading."""
    app_obj = current_app._get_current_object()
    if export_queue:
        return export_queue.enqueue(fn, app_obj, *args)
    t = threading.Thread(target=fn, args=(app_obj, *args), daemon=True)
    t.start()
    return t


def log_export(user, formato, tipo, filtros, status, message='', file_path=None):
    with app.app_context():
        entry = ExportLog(
            user=user,
            company_id=current_company_id(),
            formato=formato,
            tipo=tipo,
            filtros=json.dumps(filtros),
            status=status,
            message=message,
            file_path=file_path,
        )
        db.session.add(entry)
        db.session.commit()
        return entry.id


def _export_job(app_obj, company_id, user, start, end, estado, categoria, formato, tipo, entry_id):  # pragma: no cover - background
    """Background task that builds the export file for the given company."""
    with app_obj.app_context():
        filtros = {
            'fecha_inicio': start.strftime('%Y-%m-%d') if start else '',
            'fecha_fin': end.strftime('%Y-%m-%d') if end else '',
            'estado': estado or '',
            'categoria': categoria or '',
        }
        try:
            if os.path.isfile('maint'):
                os.remove('maint')
            os.makedirs('maint', exist_ok=True)
            q = Invoice.query.filter_by(company_id=company_id)
            if start:
                q = q.filter(Invoice.date >= start)
            if end:
                q = q.filter(Invoice.date <= end)
            if estado:
                q = q.filter(Invoice.status == estado)
            if categoria:
                q = q.join(Invoice.items).filter(InvoiceItem.category == categoria)
            path = os.path.join('maint', f'export_{entry_id}.{formato}')
            if formato == 'csv':
                with open(path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    if tipo == 'resumen':
                        writer.writerow(['Categoría', 'Cantidad', 'Total'])
                        summary = (
                            company_query(InvoiceItem)
                            .join(Invoice)
                            .with_entities(
                                InvoiceItem.category,
                                func.count(InvoiceItem.id),
                                func.sum(InvoiceItem.unit_price * InvoiceItem.quantity - InvoiceItem.discount),
                            )
                            .group_by(InvoiceItem.category)
                        )
                        if start:
                            summary = summary.filter(Invoice.date >= start)
                        if end:
                            summary = summary.filter(Invoice.date <= end)
                        if estado:
                            summary = summary.filter(Invoice.status == estado)
                        if categoria:
                            summary = summary.filter(InvoiceItem.category == categoria)
                        for cat, cnt, tot in summary:
                            writer.writerow([cat or 'Sin categoría', cnt, f"{tot or 0:.2f}"])
                    else:
                        writer.writerow(['Cliente', 'Fecha', 'Estado', 'Total'])
                        stream_q = q.options(
                            joinedload(Invoice.client),
                            load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
                        ).yield_per(100)
                        for inv in stream_q:
                            writer.writerow([
                                inv.client.name if inv.client else '',
                                inv.date.strftime('%Y-%m-%d'),
                                inv.status or '',
                                f"{inv.total:.2f}",
                            ])
            elif formato == 'xlsx' and Workbook is not None:
                wb = Workbook()
                ws = wb.active
                if tipo == 'resumen':
                    ws.append(['Categoría', 'Cantidad', 'Total'])
                    summary = (
                        company_query(InvoiceItem)
                        .join(Invoice)
                        .with_entities(
                            InvoiceItem.category,
                            func.count(InvoiceItem.id),
                            func.sum(InvoiceItem.unit_price * InvoiceItem.quantity - InvoiceItem.discount),
                        )
                        .group_by(InvoiceItem.category)
                    )
                    if start:
                        summary = summary.filter(Invoice.date >= start)
                    if end:
                        summary = summary.filter(Invoice.date <= end)
                    if estado:
                        summary = summary.filter(Invoice.status == estado)
                    if categoria:
                        summary = summary.filter(InvoiceItem.category == categoria)
                    for cat, cnt, tot in summary:
                        ws.append([cat or 'Sin categoría', cnt, float(tot or 0)])
                else:
                    ws.append(['Cliente', 'Fecha', 'Estado', 'Total'])
                    stream_q = q.options(
                        joinedload(Invoice.client),
                        load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
                    ).yield_per(100)
                    for inv in stream_q:
                        ws.append([
                            inv.client.name if inv.client else '',
                            inv.date.strftime('%Y-%m-%d'),
                            inv.status or '',
                            float(inv.total),
                        ])
                wb.save(path)
            entry = db.session.get(ExportLog, entry_id)
            entry.status = 'success'
            entry.file_path = path
            db.session.commit()
        except Exception as exc:  # pragma: no cover - hard to simulate failures
            entry = db.session.get(ExportLog, entry_id)
            entry.status = 'fail'
            entry.message = str(exc)
            db.session.commit()
def _migrate_legacy_schema():
    """Add missing columns to older SQLite databases.

    Early versions of the project lacked fields such as ``Product.category``.
    Users with an old ``database.sqlite`` would see errors like
    "no such column: product.category" when creating cotizaciones.  This helper
    checks for expected columns and adds them on the fly so the application can
    continue running without manual intervention.
    """
    inspector = inspect(db.engine)
    statements = []

    def _index_names(table_name: str) -> set[str]:
        try:
            return {idx.get('name') for idx in inspector.get_indexes(table_name) if idx.get('name')}
        except Exception:  # pragma: no cover - backend/index reflection variations
            return set()

    if inspector.has_table('product'):
        try:
            product_cols = {c['name'] for c in inspector.get_columns('product')}
        except NoSuchTableError:  # pragma: no cover - sqlite reflection race
            product_cols = set()
        if 'category' not in product_cols:
            statements.append("ALTER TABLE product ADD COLUMN category VARCHAR(50)")
        if 'unit' not in product_cols:
            statements.append("ALTER TABLE product ADD COLUMN unit VARCHAR(20) DEFAULT 'Unidad'")
        if 'has_itbis' not in product_cols:
            statements.append("ALTER TABLE product ADD COLUMN has_itbis BOOLEAN DEFAULT 1")
        if 'cost_price' not in product_cols:
            statements.append("ALTER TABLE product ADD COLUMN cost_price FLOAT")

    if not inspector.has_table('product_price_log'):
        statements.append(
            """CREATE TABLE product_price_log (
                id INTEGER PRIMARY KEY,
                product_id INTEGER NOT NULL REFERENCES product(id),
                old_price FLOAT,
                new_price FLOAT NOT NULL,
                old_cost_price FLOAT,
                new_cost_price FLOAT,
                changed_by INTEGER REFERENCES user(id),
                changed_at DATETIME NOT NULL,
                company_id INTEGER NOT NULL REFERENCES company_info(id)
            )"""
        )

    dialect_name = db.engine.dialect.name

    if inspector.has_table('user'):
        try:
            user_columns_info = inspector.get_columns('user')
            user_cols = {c['name'] for c in user_columns_info}
        except NoSuchTableError:  # pragma: no cover - sqlite reflection race
            user_columns_info = []
            user_cols = set()
        if 'email' not in user_cols:
            statements.append("ALTER TABLE user ADD COLUMN email VARCHAR(120)")
        if 'first_name' not in user_cols:
            statements.append("ALTER TABLE user ADD COLUMN first_name VARCHAR(120) DEFAULT ''")
        if 'last_name' not in user_cols:
            statements.append("ALTER TABLE user ADD COLUMN last_name VARCHAR(120) DEFAULT ''")
        if dialect_name in {'mysql', 'mariadb'}:
            pwd_col = next((c for c in user_columns_info if c['name'] == 'password'), None)
            pwd_len = getattr((pwd_col or {}).get('type'), 'length', None)
            if pwd_len and pwd_len < 255:
                statements.append("ALTER TABLE user MODIFY COLUMN password VARCHAR(255) NOT NULL")

    if inspector.has_table('account_request') and dialect_name in {'mysql', 'mariadb'}:
        try:
            req_columns_info = inspector.get_columns('account_request')
        except NoSuchTableError:  # pragma: no cover
            req_columns_info = []
        req_pwd_col = next((c for c in req_columns_info if c['name'] == 'password'), None)
        req_pwd_len = getattr((req_pwd_col or {}).get('type'), 'length', None)
        if req_pwd_len and req_pwd_len < 255:
            statements.append("ALTER TABLE account_request MODIFY COLUMN password VARCHAR(255) NOT NULL")

    if inspector.has_table('inventory_movement'):
        try:
            im_cols = {c['name'] for c in inspector.get_columns('inventory_movement')}
        except NoSuchTableError:  # pragma: no cover - sqlite reflection race
            im_cols = set()
        if 'executed_by' not in im_cols:
            statements.append(
                "ALTER TABLE inventory_movement ADD COLUMN executed_by INTEGER REFERENCES user(id)"
            )

    if inspector.has_table('quotation'):
        try:
            quote_cols = {c['name'] for c in inspector.get_columns('quotation')}
        except NoSuchTableError:  # pragma: no cover
            quote_cols = set()
        if 'status' not in quote_cols:
            statements.append("ALTER TABLE quotation ADD COLUMN status VARCHAR(20)")
        if 'valid_until' not in quote_cols:
            statements.append("ALTER TABLE quotation ADD COLUMN valid_until DATETIME")

    if not inspector.has_table('audit_log'):
        statements.append(
            """CREATE TABLE audit_log (
                id INTEGER PRIMARY KEY,
                created_at DATETIME NOT NULL,
                user_id INTEGER REFERENCES user(id),
                username VARCHAR(80),
                role VARCHAR(20),
                company_id INTEGER,
                action VARCHAR(80) NOT NULL,
                entity VARCHAR(80) NOT NULL,
                entity_id VARCHAR(80),
                status VARCHAR(20),
                details TEXT,
                ip VARCHAR(45),
                user_agent VARCHAR(255)
            )"""
        )

    if not inspector.has_table('error_report'):
        statements.append(
            """CREATE TABLE error_report (
                id INTEGER PRIMARY KEY,
                created_at DATETIME NOT NULL,
                updated_at DATETIME NOT NULL,
                user_id INTEGER REFERENCES user(id),
                username VARCHAR(80),
                company_id INTEGER,
                title VARCHAR(180) NOT NULL,
                module VARCHAR(80) NOT NULL,
                severity VARCHAR(20) NOT NULL,
                status VARCHAR(20) NOT NULL,
                page_url VARCHAR(255),
                happened_at DATETIME,
                expected_behavior TEXT,
                actual_behavior TEXT NOT NULL,
                steps_to_reproduce TEXT NOT NULL,
                contact_email VARCHAR(120),
                ip VARCHAR(45),
                user_agent VARCHAR(255),
                admin_notes TEXT
            )"""
        )

    if not inspector.has_table('system_announcement'):
        statements.append(
            """CREATE TABLE system_announcement (
                id INTEGER PRIMARY KEY,
                title VARCHAR(180) NOT NULL,
                message TEXT NOT NULL,
                scheduled_for DATETIME,
                is_active BOOLEAN NOT NULL DEFAULT 1,
                created_by INTEGER REFERENCES user(id),
                created_at DATETIME NOT NULL,
                updated_at DATETIME NOT NULL
            )"""
        )

    if not inspector.has_table('app_setting'):
        statements.append(
            """CREATE TABLE app_setting (
                `key` VARCHAR(80) PRIMARY KEY,
                value VARCHAR(255) NOT NULL,
                updated_at DATETIME NOT NULL
            )"""
        )

    if not inspector.has_table('rnc_registry'):
        statements.append(
            """CREATE TABLE rnc_registry (
                rnc VARCHAR(20) PRIMARY KEY,
                name VARCHAR(180) NOT NULL,
                source VARCHAR(40) NOT NULL,
                updated_at DATETIME NOT NULL
            )"""
        )

    if inspector.has_table('rnc_registry'):
        rnc_indexes = _index_names('rnc_registry')
        if 'ix_rnc_registry_updated_at' not in rnc_indexes:
            statements.append("CREATE INDEX ix_rnc_registry_updated_at ON rnc_registry (updated_at)")

    if inspector.has_table('audit_log'):
        audit_indexes = _index_names('audit_log')
        if 'ix_audit_log_created_at' not in audit_indexes:
            statements.append("CREATE INDEX ix_audit_log_created_at ON audit_log (created_at)")
        if 'ix_audit_log_action_created_at' not in audit_indexes:
            statements.append("CREATE INDEX ix_audit_log_action_created_at ON audit_log (action, created_at)")
        if 'ix_audit_log_entity_entity_id' not in audit_indexes:
            statements.append("CREATE INDEX ix_audit_log_entity_entity_id ON audit_log (entity, entity_id)")
        if 'ix_audit_log_user_id_created_at' not in audit_indexes:
            statements.append("CREATE INDEX ix_audit_log_user_id_created_at ON audit_log (user_id, created_at)")

    if inspector.has_table('user'):
        # Normaliza usuarios antiguos a minúsculas para evitar conflictos por mayúsculas.
        db.session.execute(db.text("UPDATE user SET username = lower(username) WHERE username <> lower(username)"))

    if inspector.has_table('account_request'):
        db.session.execute(db.text("UPDATE account_request SET username = lower(username) WHERE username <> lower(username)"))

    wrote_normalizations = False
    if inspector.has_table('user'):
        wrote_normalizations = True
    if inspector.has_table('account_request'):
        wrote_normalizations = True

    for stmt in statements:
        db.session.execute(db.text(stmt))
    if statements or wrote_normalizations:
        db.session.commit()


def run_auto_migrations():
    """Apply Alembic migrations or fallback to ``create_all``.

    This runs on import so that new fields are added automatically for
    existing SQLite databases where developers might forget to run
    ``flask db upgrade``.  It also calls :func:`_migrate_legacy_schema`
    to patch columns that predate Alembic.
    """
    with app.app_context():
        try:
            upgrade()
        except Exception:  # pragma: no cover - for environments without migrations
            db.create_all()
        _migrate_legacy_schema()


def ensure_admin():
    """Backward-compatible helper used by tests/legacy startup hooks.

    Keeps previous contract: always attempts ``upgrade()`` when invoked.
    """
    with app.app_context():
        upgrade()



# Run migrations when the module is imported so that new fields are available
# even if ``flask db upgrade`` wasn't executed manually.
run_auto_migrations()


SIGNUP_AUTO_APPROVE_KEY = 'signup_auto_approve'


def _is_signup_auto_approve_enabled() -> bool:
    setting = db.session.get(AppSetting, SIGNUP_AUTO_APPROVE_KEY)
    if not setting:
        return False
    return str(setting.value).strip().lower() in {'1', 'true', 'yes', 'on'}


def _set_signup_auto_approve(enabled: bool) -> None:
    setting = db.session.get(AppSetting, SIGNUP_AUTO_APPROVE_KEY)
    if not setting:
        setting = AppSetting(key=SIGNUP_AUTO_APPROVE_KEY, value='1' if enabled else '0')
        db.session.add(setting)
    else:
        setting.value = '1' if enabled else '0'


def _create_company_user_from_signup(*, first_name: str, last_name: str, company_name: str, identifier: str, phone: str, address: str, website: str | None, username: str, password_hash: str, email: str | None, role: str) -> tuple[CompanyInfo, User]:
    company = CompanyInfo(
        name=company_name,
        street=address or '',
        sector='',
        province='',
        phone=phone,
        rnc=identifier or '',
        website=website,
        logo='',
    )
    db.session.add(company)
    db.session.flush()
    user = User(
        username=username,
        first_name=first_name,
        last_name=last_name,
        role=role,
        company_id=company.id,
        email=email,
    )
    user.password = password_hash
    db.session.add(user)
    return company, user

# Utility constants
ITBIS_RATE = 0.18
UNITS = ('Unidad', 'Metro', 'Onza', 'Libra', 'Kilogramo', 'Litro')
CATEGORIES = (
    'Alimentos y Bebidas',
    'Productos Industriales / Materiales',
    'Minerales',
    'Salud y Cuidado Personal',
    'Electrónica y Tecnología',
    'Hogar y Construcción',
    'Energía Renovable',
    'Otros',
)
INVOICE_STATUSES = ('Pendiente', 'Pagada')
MAX_EXPORT_ROWS = 50000


QUOTATION_VALIDITY_OPTIONS = {
    '15d': 15,
    '1m': 30,
    '2m': 60,
    '3m': 90,
}


def _quotation_validity_days(value: str | None) -> int:
    return QUOTATION_VALIDITY_OPTIONS.get(value or '1m', 30)


def current_company_id():
    if not has_request_context():
        return None
    cid = session.get('company_id')
    if cid is None and session.get('role') != 'admin' and session.get('user_id'):
        user = db.session.get(User, session.get('user_id'))
        if user and user.company_id is not None:
            cid = user.company_id
            session['company_id'] = cid
    return cid


def notify(message):
    if current_company_id():
        db.session.add(Notification(company_id=current_company_id(), message=message))
        db.session.commit()


def log_audit(action, entity, entity_id=None, status='ok', details=''):
    """Persist audit trail events for CPanel review."""
    try:
        details_text = details
        if isinstance(details, (dict, list, tuple)):
            details_text = json.dumps(details, ensure_ascii=False, sort_keys=True)
        entry = AuditLog(
            user_id=session.get('user_id'),
            username=session.get('username') or session.get('full_name'),
            role=session.get('role'),
            company_id=current_company_id(),
            action=action,
            entity=entity,
            entity_id=str(entity_id) if entity_id is not None else None,
            status=status,
            details=str(details_text or ''),
            ip=(request.headers.get('X-Forwarded-For', request.remote_addr) or '')[:45],
            user_agent=(request.headers.get('User-Agent') or '')[:255],
        )
        db.session.add(entry)
        db.session.commit()
    except Exception as exc:  # pragma: no cover
        app.logger.exception('Audit log write failed: %s', exc)


def company_query(model):
    cid = current_company_id()
    if session.get('role') == 'admin' and cid is None:
        return model.query
    return model.query.filter_by(company_id=cid)


def company_get(model, object_id):
    return company_query(model).filter_by(id=object_id).first_or_404()

def _to_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _validate_product_cost_inputs(price, use_cost, cost_price_raw):
    """Validate optional product cost data and return normalized values.

    Returns a tuple of: (is_valid, cost_price_or_none, warning_message_or_none).
    """
    if not use_cost:
        return True, None, None

    cost_price = _to_float(cost_price_raw)
    if cost_price <= 0:
        return False, None, 'El costo debe ser mayor que 0 cuando activa "Usar costo".'

    margin = ((price - cost_price) / cost_price) * 100
    if price < cost_price:
        return True, cost_price, (
            'Advertencia: el precio de venta está por debajo del costo. '
            f'Margen actual: {margin:.1f}%.'
        )
    if margin < 5:
        return True, cost_price, f'Advertencia: margen bajo ({margin:.1f}%).'
    return True, cost_price, None


def _pct_change(current, previous):
    """Return percentage change or None when previous is zero/missing."""
    if previous in (None, 0):
        return None
    return ((current - previous) / previous) * 100




def _log_product_price_change(product, old_price, old_cost_price):
    """Persist a price/cost change log row when values changed."""
    if old_price == product.price and old_cost_price == product.cost_price:
        return
    db.session.add(
        ProductPriceLog(
            product_id=product.id,
            old_price=old_price,
            new_price=product.price,
            old_cost_price=old_cost_price,
            new_cost_price=product.cost_price,
            changed_by=session.get('user_id'),
            company_id=current_company_id(),
        )
    )

def generate_reference(name: str) -> str:
    """Generate a unique reference based on product name."""
    prefix = ''.join(ch for ch in (name or '').upper() if ch.isalnum())[:3]
    if not prefix:
        prefix = 'REF'
    existing = company_query(Product).filter(Product.reference.like(f"{prefix}%")).all()
    numbers = []
    for p in existing:
        if p.reference and p.reference.startswith(prefix):
            suf = p.reference[len(prefix):]
            if suf.isdigit():
                numbers.append(int(suf))
    next_no = (max(numbers) + 1) if numbers else 1
    return f"{prefix}{next_no:03d}"


def _parse_report_params(fecha_inicio, fecha_fin, estado, categoria):
    """Validate and normalize report filter parameters."""
    start = end = None
    if fecha_inicio:
        try:
            start = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        except ValueError:
            start = None
    if fecha_fin:
        try:
            end = datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            end = None
    if start and end and start > end:
        start = end = None
    if estado not in INVOICE_STATUSES:
        estado = None
    if categoria not in CATEGORIES:
        categoria = None
    return start, end, estado, categoria


@app.template_filter('phone')
def fmt_phone(value):
    digits = re.sub(r'\D', '', value or '')
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return value or ''


@app.template_filter('id_doc')
def fmt_id(value):
    digits = re.sub(r'\D', '', value or '')
    if len(digits) == 9:
        return f"{digits[:3]}-{digits[3:8]}-{digits[8:]}"
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:10]}-{digits[10:]}"
    return value or ''


def calculate_totals(items):
    subtotal = 0
    itbis = 0
    for item in items:
        line = (item['unit_price'] * item['quantity']) - item['discount']
        subtotal += line
        if item.get('has_itbis'):
            itbis += line * ITBIS_RATE
    return subtotal, itbis, subtotal + itbis


def build_items(product_ids, quantities, discounts):
    # Convert the list of product ids to integers, ignoring any non-numeric
    # values that may come from malformed form submissions. Previously a
    # stray string (e.g. the product name) would raise ``ValueError`` and the
    # quotation silently failed to save.  Now we skip those entries so the
    # view can provide proper feedback to the user.
    ids: list[int] = []
    for pid in product_ids:
        try:
            if pid:
                ids.append(int(pid))
        except (TypeError, ValueError):
            continue
    products = (
        company_query(Product)
        .options(load_only(
            Product.id,
            Product.code,
            Product.reference,
            Product.name,
            Product.unit,
            Product.price,
            Product.category,
            Product.has_itbis,
        ))
        .filter(Product.id.in_(ids))
        .all()
    )
    prod_map = {str(p.id): p for p in products}
    items = []
    for pid, q, d in zip(product_ids, quantities, discounts):
        product = prod_map.get(pid)
        if not product:
            continue
        qty = _to_int(q)
        percent = _to_float(d)
        discount_amount = product.price * qty * (percent / 100)
        items.append({
            'code': product.code,
            'reference': product.reference,
            'product_name': product.name,
            'unit': product.unit,
            'unit_price': product.price,
            'quantity': qty,
            'discount': discount_amount,
            'category': product.category,
            'has_itbis': product.has_itbis,
            'company_id': current_company_id(),
        })
    return items


@app.route('/api/rnc/<rnc>')
def rnc_lookup(rnc):
    clean = re.sub(r'\D', '', rnc or '')
    name = ''
    if clean:
        row = db.session.get(RNCRegistry, clean)
        if row:
            name = row.name
    if not name:
        name = RNC_DATA.get(clean, '')
    if not name and clean:
        client = Client.query.filter(func.replace(Client.identifier, '-', '') == clean).first()
        name = client.name if client else ''
    return jsonify({'name': name})


@app.before_request
def load_company():
    cid = current_company_id()
    g.company = db.session.get(CompanyInfo, cid) if cid else None


@app.context_processor
def inject_company():
    notif_count = 0
    active_announcement = None
    try:
        cid = current_company_id()
        if 'user_id' in session and cid:
            # Avoid expensive inventory scans on every request.
            # Refresh low-stock notifications at most once every 5 minutes per session.
            now_ts = int(time.time())
            last_scan = session.get('low_stock_scan_at', 0)
            if now_ts - last_scan >= 300:
                low_stock = (
                    company_query(ProductStock)
                    .filter(ProductStock.stock <= ProductStock.min_stock, ProductStock.min_stock > 0)
                    .all()
                )
                for ps in low_stock:
                    msg = f"Stock bajo: {ps.product.name}"
                    exists = Notification.query.filter_by(company_id=cid, message=msg).first()
                    if not exists:
                        db.session.add(Notification(company_id=cid, message=msg))
                if low_stock:
                    db.session.commit()
                session['low_stock_scan_at'] = now_ts

            notif_count = Notification.query.filter_by(company_id=cid, is_read=False).count()
        active_announcement = (
            SystemAnnouncement.query
            .filter_by(is_active=True)
            .order_by(SystemAnnouncement.updated_at.desc())
            .first()
        )
    except Exception as exc:
        app.logger.exception('Failed to compute notifications: %s', exc)
    return {
        'company': getattr(g, 'company', None),
        'notification_count': notif_count,
        'current_dom_time': dom_now().strftime('%d/%m/%Y %I:%M %p'),
        'active_announcement': active_announcement,
        'app_version': APP_VERSION,
        'app_version_highlights': APP_VERSION_HIGHLIGHTS,
    }


def get_company_info():
    c = db.session.get(CompanyInfo, current_company_id())
    if not c:
        return {}
    return {
        'name': c.name,
        'address': f"{c.street}, {c.sector}, {c.province}",
        'rnc': c.rnc,
        'phone': c.phone,
        'website': c.website,
        'logo': os.path.join(app.static_folder, c.logo if c.logo.startswith('uploads/') else f'uploads/{c.logo}') if c.logo else None,
        'ncf_final': c.ncf_final,
        'ncf_fiscal': c.ncf_fiscal,
    }


def _company_short_slug(company_name: str | None) -> str:
    base = (company_name or 'empresa').strip().lower()
    base = re.sub(r'[^a-z0-9\s-]', '', base)
    base = re.sub(r'[\s_-]+', '-', base).strip('-')
    return (base[:40] or 'empresa')


def _company_private_token(company_id: int | None, company_name: str | None) -> str:
    seed = f"{app.config.get('SECRET_KEY','tiendix')}:{company_id or 0}:{company_name or ''}"
    token_number = int(hashlib.sha256(seed.encode('utf-8')).hexdigest(), 16) % 1_000_000
    return f"{token_number:06d}"


def _public_doc_url(doc_type: str, doc_number: int | str, *, company_name: str | None = None, company_id: int | None = None) -> str | None:
    base_url = (app.config.get('PUBLIC_DOCS_BASE_URL') or '').strip().rstrip('/')
    if not base_url:
        return None
    cid = company_id if company_id is not None else current_company_id()
    name = company_name or (getattr(g, 'company', None).name if getattr(g, 'company', None) else None)
    short = _company_short_slug(name)
    token = _company_private_token(cid, name)
    safe_type = secure_filename((doc_type or 'documento').lower()) or 'documento'
    number = f"{int(doc_number):02d}" if str(doc_number).isdigit() else secure_filename(str(doc_number))
    return f"{base_url}/{short}/{token}/{safe_type}/{number}.pdf"


def _archive_root_dir() -> Path:
    configured = (app.config.get('PDF_ARCHIVE_ROOT') or '').strip()
    if configured:
        return Path(configured)
    return Path(app.root_path) / 'generated_docs'


def _archived_pdf_path(doc_type: str, doc_number: int | str, *, company_name: str | None = None, company_id: int | None = None) -> Path:
    cid = company_id if company_id is not None else current_company_id()
    name = company_name or (getattr(g, 'company', None).name if getattr(g, 'company', None) else None)
    short = _company_short_slug(name)
    token = _company_private_token(cid, name)
    safe_type = secure_filename((doc_type or 'documento').lower()) or 'documento'
    number = f"{int(doc_number):02d}" if str(doc_number).isdigit() else secure_filename(str(doc_number))
    return _archive_root_dir() / short / token / safe_type / f"{number}.pdf"


def _build_quotation_pdf_bytes(quotation: Quotation, company: dict[str, str | None]) -> bytes:
    return generate_pdf_bytes(
        'Cotización',
        company,
        quotation.client,
        quotation.items,
        quotation.subtotal,
        quotation.itbis,
        quotation.total,
        seller=quotation.seller,
        payment_method=quotation.payment_method,
        bank=quotation.bank,
        doc_number=quotation.id,
        note=quotation.note,
        date=quotation.date,
        valid_until=quotation.valid_until,
        footer=(
            "Condiciones: Esta cotización es válida por 30 días a partir de la fecha de emisión. "
            "Los precios están sujetos a cambios sin previo aviso. "
            "El ITBIS ha sido calculado conforme a la ley vigente."
        ),
    )


def _build_order_pdf_bytes(order: Order, company: dict[str, str | None]) -> bytes:
    return generate_pdf_bytes(
        'Pedido',
        company,
        order.client,
        order.items,
        order.subtotal,
        order.itbis,
        order.total,
        seller=order.seller,
        payment_method=order.payment_method,
        bank=order.bank,
        doc_number=order.id,
        note=order.note,
        date=order.date,
        footer=(
            "Este pedido será procesado tras la confirmación de pago. "
            "Tiempo estimado de entrega: 3 a 5 días hábiles."
        ),
    )


def _build_invoice_pdf_bytes(invoice: Invoice, company: dict[str, str | None]) -> bytes:
    return generate_pdf_bytes(
        'Factura',
        company,
        invoice.client,
        invoice.items,
        invoice.subtotal,
        invoice.itbis,
        invoice.total,
        ncf=invoice.ncf,
        seller=invoice.seller,
        payment_method=invoice.payment_method,
        bank=invoice.bank,
        purchase_order=invoice.order.customer_po if invoice.order else None,
        doc_number=invoice.id,
        invoice_type=invoice.invoice_type,
        note=invoice.note,
        date=invoice.date,
        footer=(
            "Factura generada electrónicamente, válida sin firma ni sello. "
            "Para reclamaciones favor comunicarse dentro de las 48 horas siguientes a la emisión. "
            "Gracias por su preferencia."
        ),
    )


def ensure_pdf_archive_environment() -> None:
    # Best-effort: ensure archive root and document-type folders exist.
    try:
        root = _archive_root_dir()
        root.mkdir(parents=True, exist_ok=True)
        for company in CompanyInfo.query.all():
            short = _company_short_slug(company.name)
            token = _company_private_token(company.id, company.name)
            base = root / short / token
            for doc_type in ('cotizacion', 'pedido', 'factura', 'estado_cuenta', 'reporte'):
                (base / doc_type).mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        app.logger.warning('Could not pre-create PDF archive directories: %s', exc)


with app.app_context():
    ensure_pdf_archive_environment()


def _archive_pdf_copy(doc_type: str, doc_number: int | str, pdf_data: bytes, company_name: str | None = None, company_id: int | None = None) -> str | None:
    # Create a cPanel-visible archive copy, without breaking download on failure.
    try:
        out_path = _archived_pdf_path(
            doc_type,
            doc_number,
            company_name=company_name,
            company_id=company_id,
        )
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(pdf_data)
        return str(out_path)
    except Exception as exc:  # pragma: no cover
        app.logger.warning('Could not archive PDF copy (%s %s): %s', doc_type, doc_number, exc)
        return None



def _archive_and_send_pdf(*, doc_type: str, doc_number: int | str, pdf_data: bytes, download_name: str, company_name: str | None = None, archive: bool = True):
    """Archive PDF copy (best-effort) and return download response.

    Uses a compatibility fallback for older Flask/Werkzeug versions that still
    expect ``attachment_filename`` instead of ``download_name``.
    """
    if archive:
        _archive_pdf_copy(doc_type, doc_number, pdf_data, company_name=company_name)
    payload = BytesIO(pdf_data)
    try:
        return send_file(
            payload,
            download_name=download_name,
            mimetype='application/pdf',
            as_attachment=True,
        )
    except TypeError:
        payload.seek(0)
        return send_file(
            payload,
            attachment_filename=download_name,
            mimetype='application/pdf',
            as_attachment=True,
        )
# Routes
@app.before_request
def require_login():
    allowed = {
        'auth.login',
        'static',
        'request_account',
        'auth.logout',
        'auth.reset_request',
        'auth.reset_password',
        'terminos',
    }
    if request.endpoint not in allowed and 'user_id' not in session:
        return redirect(url_for('auth.login'))
    admin_extra = {'admin_companies', 'select_company', 'clear_company',
                   'admin_requests', 'approve_request', 'reject_request'}
    if session.get('role') == 'admin' and not session.get('company_id') \
            and request.endpoint not in allowed.union(admin_extra):
        return redirect(url_for('admin_companies'))


def admin_only(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Acceso restringido')
            return redirect(url_for('list_quotations'))
        return f(*args, **kwargs)
    return wrapper


def manager_only(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get('role') not in ('admin', 'manager'):
            flash('Acceso restringido')
            return redirect(url_for('list_quotations'))
        return f(*args, **kwargs)
    return wrapper

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    return redirect(url_for('list_quotations'))




def _build_signup_captcha() -> tuple[str, str]:
    a = random.randint(2, 9)
    b = random.randint(1, 9)
    return f"¿Cuánto es {a} + {b}?", str(a + b)

@app.route('/solicitar-cuenta', methods=['GET', 'POST'])
def request_account():
    form = AccountRequestForm()

    if request.method == 'GET' or 'signup_captcha_answer' not in session:
        q, answer = _build_signup_captcha()
        session['signup_captcha_question'] = q
        session['signup_captcha_answer'] = answer

    if form.validate_on_submit():
        if not form.accepted_terms.data:
            flash(
                'Debe aceptar los Términos y Condiciones para crear una cuenta en Tiendix.',
                'request',
            )
            return redirect(url_for('request_account'))
        password = (request.form.get('password') or '')
        if len(password) < 6:
            flash('La contraseña debe tener mínimo 6 caracteres', 'request')
            return redirect(url_for('request_account'))
        if password != request.form.get('confirm_password'):
            flash('Las contraseñas no coinciden', 'request')
            return redirect(url_for('request_account'))
        username = (request.form.get('username') or '').strip().lower()
        if not username:
            flash('Debe ingresar un usuario válido', 'request')
            return redirect(url_for('request_account'))

        required_labels = {
            'first_name': 'Nombre',
            'last_name': 'Apellido',
            'phone': 'Teléfono',
            'company': 'Empresa o marca',
            'identifier': 'Cédula/RNC',
            'email': 'Correo',
            'username': 'Usuario',
        }
        missing = [label for key, label in required_labels.items() if not (request.form.get(key) or '').strip()]
        if missing:
            flash(f"Complete los campos requeridos: {', '.join(missing)}", 'request')
            return redirect(url_for('request_account'))

        if not app.config.get('TESTING', False):
            expected = str(session.get('signup_captcha_answer') or '').strip()
            got = (request.form.get('captcha_answer') or '').strip()
            if not expected or got != expected:
                flash('Captcha inválido. Intente nuevamente.', 'request')
                q, answer = _build_signup_captcha()
                session['signup_captcha_question'] = q
                session['signup_captcha_answer'] = answer
                return redirect(url_for('request_account'))
        existing_user = User.query.filter(func.lower(User.username) == username).first()
        pending_user = AccountRequest.query.filter(func.lower(AccountRequest.username) == username).first()
        if existing_user or pending_user:
            flash('Ese nombre de usuario ya existe, use otro.', 'request')
            return redirect(url_for('request_account'))
        account_type = request.form['account_type']
        identifier = request.form.get('identifier')
        if not identifier:
            flash('Debe ingresar RNC o Cédula', 'request')
            return redirect(url_for('request_account'))
        if User.query.count() == 0:
            company, owner = _create_company_user_from_signup(
                first_name=request.form['first_name'],
                last_name=request.form['last_name'],
                company_name=request.form['company'],
                identifier=identifier,
                phone=request.form['phone'],
                address=request.form.get('address') or '',
                website=request.form.get('website'),
                username=username,
                password_hash=generate_password_hash(password),
                email=request.form['email'],
                role='admin',
            )
            db.session.commit()
            log_audit('owner_bootstrap', 'user', owner.id, details=f'company={company.id};username={username}')
            flash('Cuenta principal creada: ahora eres Administrador (dueño).', 'login')
            session.pop('signup_captcha_answer', None)
            session.pop('signup_captcha_question', None)
            return redirect(url_for('auth.login'))

        if _is_signup_auto_approve_enabled():
            company, user = _create_company_user_from_signup(
                first_name=request.form['first_name'],
                last_name=request.form['last_name'],
                company_name=request.form['company'],
                identifier=identifier,
                phone=request.form['phone'],
                address=request.form.get('address') or '',
                website=request.form.get('website'),
                username=username,
                password_hash=generate_password_hash(password),
                email=request.form['email'],
                role='manager',
            )
            db.session.commit()
            log_audit('account_auto_approved', 'user', user.id, details=f'company={company.id};username={username}')
            flash('Cuenta creada correctamente con rol Manager. Ya puede iniciar sesión.', 'login')
            session.pop('signup_captcha_answer', None)
            session.pop('signup_captcha_question', None)
            return redirect(url_for('auth.login'))

        req = AccountRequest(
            account_type=account_type,
            first_name=request.form['first_name'],
            last_name=request.form['last_name'],
            company=request.form['company'],
            identifier=identifier,
            phone=request.form['phone'],
            email=request.form['email'],
            address=request.form.get('address'),
            website=request.form.get('website'),
            username=username,
            password=generate_password_hash(password),
            accepted_terms=True,
            accepted_terms_at=dom_now(),
            accepted_terms_ip=request.remote_addr,
            accepted_terms_user_agent=request.headers.get('User-Agent', ''),
        )
        db.session.add(req)
        db.session.commit()
        log_audit('account_request_create', 'account_request', req.id, details=f'email={req.email};company={req.company}')
        flash('Solicitud enviada, espere aprobación', 'login')
        session.pop('signup_captcha_answer', None)
        session.pop('signup_captcha_question', None)
        return redirect(url_for('auth.login'))
    elif request.method == 'POST':
        flash(
            'Debe aceptar los Términos y Condiciones para crear una cuenta en Tiendix.',
            'request',
        )
        return redirect(url_for('request_account'))
    return render_template('solicitar_cuenta.html', form=form, captcha_question=session.get('signup_captcha_question', ''))


@app.route('/terminos')
def terminos():
    return render_template('terminos.html')


@app.route('/admin/solicitudes')
@admin_only
def admin_requests():
    requests = AccountRequest.query.all()
    return render_template('admin_solicitudes.html', requests=requests, signup_auto_approve=_is_signup_auto_approve_enabled())


@app.route('/admin/companies')
@admin_only
def admin_companies():
    companies = CompanyInfo.query.all()
    return render_template('admin_companies.html', companies=companies)


@app.route('/admin/companies/select/<int:company_id>')
@admin_only
def select_company(company_id):
    session['company_id'] = company_id
    return redirect(url_for('list_quotations'))


@app.route('/admin/companies/clear')
@admin_only
def clear_company():
    session.pop('company_id', None)
    return redirect(url_for('admin_companies'))


@app.route('/admin/solicitudes/<int:req_id>/aprobar', methods=['POST'])
@admin_only
def approve_request(req_id):
    req = AccountRequest.query.get_or_404(req_id)
    role = request.form.get('role', 'company')
    username = (req.username or '').lower()
    if User.query.filter(func.lower(User.username) == username).first():
        flash('No se puede aprobar: el usuario ya existe (mayúsculas/minúsculas).')
        return redirect(url_for('admin_requests'))
    password = req.password
    email = req.email
    company, user = _create_company_user_from_signup(
        first_name=req.first_name,
        last_name=req.last_name,
        company_name=req.company,
        identifier=req.identifier or '',
        phone=req.phone,
        address=req.address or '',
        website=req.website,
        username=username,
        password_hash=password,
        email=email,
        role=role,
    )
    db.session.delete(req)
    db.session.commit()

    # Envío de enlace temporal para establecer o restablecer contraseña
    token = generate_reset_token(user)
    html = render_template(
        'emails/account_approved.html',
        username=username,
        company=company.name,
        login_url=url_for('auth.login', _external=True),
        reset_url=url_for('auth.reset_password', token=token, _external=True),
    )
    send_email(email, 'Tu cuenta ha sido aprobada', html)
    log_audit('account_request_approve', 'account_request', req_id, details=f'user={user.id};company={company.id}')
    flash('Cuenta aprobada')
    return redirect(url_for('admin_requests'))


@app.route('/admin/solicitudes/<int:req_id>/rechazar', methods=['POST'])
@admin_only
def reject_request(req_id):
    req = AccountRequest.query.get_or_404(req_id)
    db.session.delete(req)
    db.session.commit()
    log_audit('account_request_reject', 'account_request', req.id)
    flash('Solicitud rechazada')
    return redirect(url_for('admin_requests'))




@app.route('/reportar-error', methods=['GET', 'POST'])
@app.route('/reportar-problema', methods=['GET', 'POST'])
def report_error():
    if 'user_id' not in session:
        flash('Debe iniciar sesión para reportar un error')
        return redirect(url_for('auth.login'))

    modules = [
        'Login / Acceso',
        'Cotizaciones',
        'Pedidos',
        'Facturas',
        'Inventario',
        'Clientes',
        'Reportes',
        'PDF / Descargas',
        'Configuración',
        'Otro',
    ]
    severities = [('baja', 'Baja'), ('media', 'Media'), ('alta', 'Alta'), ('critica', 'Crítica')]

    if request.method == 'POST':
        title = (request.form.get('title') or '').strip()
        module = request.form.get('module') or 'Otro'
        severity = request.form.get('severity') or 'media'
        page_url = (request.form.get('page_url') or '').strip()
        expected_behavior = (request.form.get('expected_behavior') or '').strip()
        actual_behavior = (request.form.get('actual_behavior') or '').strip()
        steps_to_reproduce = (request.form.get('steps_to_reproduce') or '').strip()
        contact_email = (request.form.get('contact_email') or '').strip()
        happened_at_raw = (request.form.get('happened_at') or '').strip()

        if not title or not actual_behavior or not steps_to_reproduce:
            flash('Complete título, qué pasó y pasos para reproducir el error.')
            return render_template('report_error.html', modules=modules, severities=severities)

        happened_at = None
        if happened_at_raw:
            try:
                happened_at = datetime.fromisoformat(happened_at_raw)
            except ValueError:
                flash('La fecha/hora del incidente no es válida.')
                return render_template('report_error.html', modules=modules, severities=severities)

        report = ErrorReport(
            user_id=session.get('user_id'),
            username=session.get('username') or session.get('full_name'),
            company_id=current_company_id(),
            title=title,
            module=module,
            severity=severity,
            status='abierto',
            page_url=page_url[:255] if page_url else None,
            happened_at=happened_at,
            expected_behavior=expected_behavior,
            actual_behavior=actual_behavior,
            steps_to_reproduce=steps_to_reproduce,
            contact_email=contact_email[:120] if contact_email else None,
            ip=(request.headers.get('X-Forwarded-For', request.remote_addr) or '')[:45],
            user_agent=(request.headers.get('User-Agent') or '')[:255],
        )
        db.session.add(report)
        db.session.commit()
        log_audit('error_report_create', 'error_report', report.id, details=f'severity={severity};module={module}')
        flash('Reporte enviado. Responderemos en un plazo de hasta 72 horas laborables (puede ser antes).')
        return redirect(url_for('report_error'))

    return render_template('report_error.html', modules=modules, severities=severities)

# --- CPanel ---


@app.route('/cpaneltx')
@admin_only
def cpanel_home():
    return render_template('cpaneltx.html', signup_auto_approve=_is_signup_auto_approve_enabled())


@app.post('/cpaneltx/signup-mode')
@admin_only
def cpanel_signup_mode():
    enabled = bool(request.form.get('signup_auto_approve'))
    _set_signup_auto_approve(enabled)
    db.session.commit()
    log_audit('cpanel_signup_mode', 'app_setting', SIGNUP_AUTO_APPROVE_KEY, details=f'enabled={enabled}')
    if enabled:
        flash('Aprobación automática activada: nuevas cuentas se crearán directamente con rol Manager.')
    else:
        flash('Aprobación automática desactivada: las solicitudes requerirán aprobación de administrador.')
    return redirect(url_for('cpanel_home'))


@app.route('/cpaneltx/rnc', methods=['GET', 'POST'])
@admin_only
def cpanel_rnc_import():
    if request.method == 'POST':
        file = request.files.get('rnc_file')
        if not file or not file.filename:
            flash('Debe seleccionar un archivo RNC .txt')
            return redirect(url_for('cpanel_rnc_import'))

        filename = (file.filename or '').lower()
        if not filename.endswith('.txt'):
            flash('Formato inválido. Debe subir un archivo .txt')
            return redirect(url_for('cpanel_rnc_import'))

        inserted = 0
        updated = 0
        skipped = 0

        parsed_rows: dict[str, str] = {}
        for raw_bytes in file.stream:
            raw_line = raw_bytes.decode('utf-8', errors='ignore')
            rnc, name = _parse_rnc_line(raw_line)
            if not rnc or not name:
                skipped += 1
                continue

            # Si hay RNC repetido en el TXT, prevalece la última ocurrencia.
            parsed_rows[rnc] = name

        if not parsed_rows:
            flash('No se encontraron registros válidos en el archivo.')
            return redirect(url_for('cpanel_rnc_import'))

        existing_rows = {
            row.rnc: row
            for row in RNCRegistry.query.filter(RNCRegistry.rnc.in_(list(parsed_rows.keys()))).all()
        }

        for rnc, name in parsed_rows.items():
            current = existing_rows.get(rnc)
            if current:
                if current.name != name:
                    current.name = name
                    current.source = 'cpanel_upload'
                    updated += 1
                RNC_DATA[rnc] = name
            else:
                db.session.add(RNCRegistry(rnc=rnc, name=name, source='cpanel_upload'))
                RNC_DATA[rnc] = name
                inserted += 1

        db.session.commit()
        log_audit('cpanel_rnc_import', 'rnc_registry', status='ok', details={
            'filename': file.filename,
            'inserted': inserted,
            'updated': updated,
            'skipped': skipped,
            'total_registry': db.session.query(func.count(RNCRegistry.rnc)).scalar(),
        })
        flash(f'Importación completada: nuevos={inserted}, actualizados={updated}, omitidos={skipped}.')
        return redirect(url_for('cpanel_rnc_import'))

    total = db.session.query(func.count(RNCRegistry.rnc)).scalar() or 0
    latest = RNCRegistry.query.order_by(RNCRegistry.updated_at.desc()).limit(20).all()
    return render_template('cpanel_rnc_import.html', total=total, latest=latest)


@app.route('/cpaneltx/avisos', methods=['GET', 'POST'])
@admin_only
def cpanel_announcements():
    if request.method == 'POST':
        title = (request.form.get('title') or '').strip()
        message = (request.form.get('message') or '').strip()
        scheduled_for_raw = (request.form.get('scheduled_for') or '').strip()
        is_active = request.form.get('is_active') == 'on'

        if not title or not message:
            flash('Título y mensaje son obligatorios')
            return redirect(url_for('cpanel_announcements'))

        scheduled_for = None
        if scheduled_for_raw:
            try:
                scheduled_for = datetime.fromisoformat(scheduled_for_raw)
            except ValueError:
                flash('La fecha/hora programada no es válida')
                return redirect(url_for('cpanel_announcements'))

        ann = SystemAnnouncement(
            title=title,
            message=message,
            scheduled_for=scheduled_for,
            is_active=is_active,
            created_by=session.get('user_id'),
        )
        db.session.add(ann)
        db.session.commit()
        log_audit('announcement_create', 'system_announcement', ann.id, details=f'active={is_active}')
        flash('Aviso general creado')
        return redirect(url_for('cpanel_announcements'))

    announcements = SystemAnnouncement.query.order_by(SystemAnnouncement.updated_at.desc()).all()
    return render_template('cpanel_announcements.html', announcements=announcements)


@app.post('/cpaneltx/avisos/<int:ann_id>/estado')
@admin_only
def cpanel_announcement_status(ann_id):
    ann = SystemAnnouncement.query.get_or_404(ann_id)
    ann.is_active = request.form.get('is_active') == 'on'
    db.session.commit()
    log_audit('announcement_status', 'system_announcement', ann.id, details=f'active={ann.is_active}')
    flash('Estado de aviso actualizado')
    return redirect(url_for('cpanel_announcements'))


@app.route('/cpaneltx/users')
@admin_only
def cpanel_users():
    q = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)
    query = User.query
    if q:
        query = query.filter(User.username.ilike(f'%{q}%'))
    users = query.order_by(User.id).paginate(page=page, per_page=10, error_out=False)
    return render_template('cpanel_users.html', users=users, q=q)


@app.post('/cpaneltx/users/<int:user_id>/update')
@admin_only
def cpanel_user_update(user_id):
    user = User.query.get_or_404(user_id)
    email = (request.form.get('email') or '').strip()
    first_name = (request.form.get('first_name') or '').strip()
    last_name = (request.form.get('last_name') or '').strip()
    password = request.form.get('password') or ''

    changed = {}

    if first_name and first_name != (user.first_name or ''):
        changed['first_name'] = {'from': user.first_name or '', 'to': first_name}
        user.first_name = first_name

    if last_name and last_name != (user.last_name or ''):
        changed['last_name'] = {'from': user.last_name or '', 'to': last_name}
        user.last_name = last_name

    if email:
        if email != (user.email or ''):
            changed['email'] = {'from': user.email or '', 'to': email}
        user.email = email

    if password:
        if len(password) < 6:
            flash('La contraseña debe tener mínimo 6 caracteres')
            return redirect(url_for('cpanel_users'))
        user.set_password(password)
        changed['password'] = {'changed': True, 'length': len(password)}

    if not changed:
        flash('No se detectaron cambios para este usuario')
        return redirect(url_for('cpanel_users'))

    db.session.commit()
    log_audit('cpanel_user_update', 'user', user_id, details={
        'target': {
            'id': user.id,
            'username': user.username,
            'company_id': user.company_id,
            'role': user.role,
        },
        'changes': changed,
    })
    flash('Usuario actualizado')
    return redirect(url_for('cpanel_users'))


@app.post('/cpaneltx/users/<int:user_id>/role')
@admin_only
def cpanel_user_role(user_id):
    user = User.query.get_or_404(user_id)
    role = request.form.get('role')
    if role in ('admin', 'manager', 'company'):
        previous_role = user.role
        if previous_role == role:
            flash('El usuario ya tenía ese rol')
            return redirect(url_for('cpanel_users'))
        user.role = role
        db.session.commit()
        flash('Rol actualizado')
        log_audit('cpanel_user_role', 'user', user_id, details={
            'target': {
                'id': user.id,
                'username': user.username,
                'company_id': user.company_id,
            },
            'role': {
                'from': previous_role,
                'to': role,
            },
        })
    return redirect(url_for('cpanel_users'))


@app.post('/cpaneltx/users/<int:user_id>/delete')
@admin_only
def cpanel_user_delete(user_id):
    user = User.query.get_or_404(user_id)
    deleted_snapshot = {
        'id': user.id,
        'username': user.username,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'email': user.email,
        'role': user.role,
        'company_id': user.company_id,
    }
    db.session.delete(user)
    db.session.commit()
    flash('Usuario eliminado')
    log_audit('cpanel_user_delete', 'user', user_id, details={'deleted_user': deleted_snapshot})
    return redirect(url_for('cpanel_users'))


@app.route('/cpaneltx/users/<int:user_id>/actividad')
@admin_only
def cpanel_user_activity(user_id):
    user = User.query.get_or_404(user_id)
    page = request.args.get('page', 1, type=int)
    logs = (
        AuditLog.query
        .filter(
            or_(
                and_(AuditLog.entity == 'user', AuditLog.entity_id == str(user_id)),
                AuditLog.user_id == user_id,
            )
        )
        .order_by(AuditLog.created_at.desc())
        .paginate(page=page, per_page=25, error_out=False)
    )
    return render_template('cpanel_user_activity.html', target_user=user, logs=logs)


@app.route('/cpaneltx/companies')
@admin_only
def cpanel_companies():
    q = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)
    query = CompanyInfo.query
    if q:
        query = query.filter(CompanyInfo.name.ilike(f'%{q}%'))
    companies = query.order_by(CompanyInfo.id).paginate(page=page, per_page=10, error_out=False)
    return render_template('cpanel_companies.html', companies=companies, q=q)




@app.get('/cpaneltx/companies/select/<int:company_id>')
@admin_only
def cpanel_company_select(company_id):
    company = db.session.get(CompanyInfo, company_id)
    if not company:
        flash('Empresa no encontrada')
        return redirect(url_for('cpanel_companies'))
    session['company_id'] = company_id
    flash(f'Ahora estás administrando la empresa: {company.name}')
    return redirect(url_for('list_quotations'))


@app.get('/cpaneltx/companies/clear')
@admin_only
def cpanel_company_clear():
    session.pop('company_id', None)
    flash('Vista global de administrador restaurada')
    return redirect(url_for('cpanel_companies'))

@app.post('/cpaneltx/companies/<int:cid>/delete')
@admin_only
def cpanel_company_delete(cid):
    company = CompanyInfo.query.get_or_404(cid)
    db.session.delete(company)
    db.session.commit()
    flash('Empresa eliminada')
    log_audit('cpanel_company_delete', 'company', cid)
    return redirect(url_for('cpanel_companies'))


@app.route('/cpaneltx/orders')
@admin_only
def cpanel_orders():
    orders = Order.query.options(joinedload(Order.client)).all()
    return render_template('cpanel_orders.html', orders=orders)


@app.post('/cpaneltx/orders/<int:oid>/delete')
@admin_only
def cpanel_order_delete(oid):
    order = Order.query.get_or_404(oid)
    db.session.delete(order)
    db.session.commit()
    flash('Pedido eliminado')
    log_audit('cpanel_order_delete', 'order', oid)
    return redirect(url_for('cpanel_orders'))


@app.route('/cpaneltx/invoices')
@admin_only
def cpanel_invoices():
    invoices = Invoice.query.options(joinedload(Invoice.client)).all()
    return render_template('cpanel_invoices.html', invoices=invoices)


@app.post('/cpaneltx/invoices/<int:iid>/delete')
@admin_only
def cpanel_invoice_delete(iid):
    inv = Invoice.query.get_or_404(iid)
    db.session.delete(inv)
    db.session.commit()
    flash('Factura eliminada')
    log_audit('cpanel_invoice_delete', 'invoice', iid)
    return redirect(url_for('cpanel_invoices'))





@app.route('/cpaneltx/reportes-error')
@admin_only
def cpanel_error_reports():
    status = request.args.get('status', '')
    severity = request.args.get('severity', '')
    q = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)

    query = ErrorReport.query
    if status:
        query = query.filter(ErrorReport.status == status)
    if severity:
        query = query.filter(ErrorReport.severity == severity)
    if q:
        query = query.filter(or_(ErrorReport.title.ilike(f'%{q}%'), ErrorReport.username.ilike(f'%{q}%'), ErrorReport.ip.ilike(f'%{q}%')))

    reports = query.order_by(ErrorReport.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('cpanel_error_reports.html', reports=reports, status=status, severity=severity, q=q)


@app.post('/cpaneltx/reportes-error/<int:report_id>/estado')
@admin_only
def cpanel_error_report_status(report_id):
    report = ErrorReport.query.get_or_404(report_id)
    status = request.form.get('status')
    if status in ('abierto', 'en_proceso', 'resuelto', 'cerrado'):
        report.status = status
        report.admin_notes = (request.form.get('admin_notes') or '').strip()
        db.session.commit()
        log_audit('error_report_status', 'error_report', report.id, details=f'status={status}')
        flash('Estado del reporte actualizado')
    else:
        flash('Estado inválido')
    return redirect(url_for('cpanel_error_reports'))

@app.route('/cpaneltx/auditoria')
@admin_only
def cpanel_auditoria():
    action = request.args.get('action', '')
    status = request.args.get('status', '')
    ip = request.args.get('ip', '')
    user_q = request.args.get('user', '')
    page = request.args.get('page', 1, type=int)

    q = AuditLog.query
    if action:
        q = q.filter(AuditLog.action.ilike(f'%{action}%'))
    if status:
        q = q.filter(AuditLog.status == status)
    if ip:
        q = q.filter(AuditLog.ip.ilike(f'%{ip}%'))
    if user_q:
        q = q.filter(AuditLog.username.ilike(f'%{user_q}%'))

    logs = q.order_by(AuditLog.created_at.desc()).paginate(page=page, per_page=30, error_out=False)

    suspicious_ips = (
        db.session.query(AuditLog.ip, func.count(AuditLog.id))
        .filter(AuditLog.action == 'account_request_create')
        .group_by(AuditLog.ip)
        .having(func.count(AuditLog.id) >= 2)
        .order_by(func.count(AuditLog.id).desc())
        .limit(20)
        .all()
    )

    return render_template(
        'cpanel_auditoria.html',
        logs=logs,
        action=action,
        status=status,
        ip=ip,
        user=user_q,
        suspicious_ips=suspicious_ips,
    )


# Clients CRUD
@app.route('/clientes', methods=['GET', 'POST'])
def clients():
    if request.method == 'POST':
        is_final = request.form.get('type') == 'final'
        identifier = request.form.get('identifier') if not is_final else request.form.get('identifier') or None
        last_name = request.form.get('last_name') if is_final else None
        if not is_final and not identifier:
            flash('El RNC es obligatorio para empresas')
            return redirect(url_for('clients'))
        if identifier:
            exists = company_query(Client).filter(Client.identifier == identifier).first()
            if exists:
                flash('Ya existe un cliente con ese RNC/Cédula')
                return redirect(url_for('clients'))
        email = request.form.get('email')
        if email:
            exists = company_query(Client).filter(Client.email == email).first()
            if exists:
                flash('Ya existe un cliente con ese correo electrónico')
                return redirect(url_for('clients'))
        client = Client(
            name=request.form['name'],
            last_name=last_name,
            identifier=identifier,
            phone=request.form.get('phone'),
            email=request.form.get('email'),
            street=request.form.get('street'),
            sector=request.form.get('sector'),
            province=request.form.get('province'),
            is_final_consumer=is_final,
            company_id=current_company_id()
        )
        db.session.add(client)
        db.session.commit()
        flash('Cliente agregado')
        notify('Cliente agregado')
        return redirect(url_for('clients'))
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    query = company_query(Client)
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Client.name.ilike(like),
                Client.last_name.ilike(like),
                Client.identifier.ilike(like),
                Client.email.ilike(like),
            )
        )
    clients = query.order_by(Client.id).paginate(page=page, per_page=25, error_out=False)
    return render_template('clientes.html', clients=clients, q=q)

@app.route('/clientes/delete/<int:client_id>', methods=['POST'])
def delete_client(client_id):
    client = company_get(Client, client_id)
    db.session.delete(client)
    db.session.commit()
    flash('Cliente eliminado')
    return redirect(url_for('clients'))

@app.route('/clientes/edit/<int:client_id>', methods=['GET', 'POST'])
def edit_client(client_id):
    client = company_get(Client, client_id)
    if request.method == 'POST':
        is_final = request.form.get('type') == 'final'
        identifier = request.form.get('identifier') if not is_final else request.form.get('identifier') or None
        last_name = request.form.get('last_name') if is_final else None
        if not is_final and not identifier:
            flash('El RNC es obligatorio para empresas')
            return redirect(url_for('edit_client', client_id=client.id))
        if identifier:
            exists = company_query(Client).filter(
                Client.identifier == identifier, Client.id != client.id
            ).first()
            if exists:
                flash('Ya existe un cliente con ese RNC/Cédula')
                return redirect(url_for('edit_client', client_id=client.id))
        email = request.form.get('email')
        if email:
            exists = company_query(Client).filter(
                Client.email == email, Client.id != client.id
            ).first()
            if exists:
                flash('Ya existe un cliente con ese correo electrónico')
                return redirect(url_for('edit_client', client_id=client.id))
        client.name = request.form['name']
        client.last_name = last_name
        client.identifier = identifier
        client.phone = request.form.get('phone')
        client.email = request.form.get('email')
        client.street = request.form.get('street')
        client.sector = request.form.get('sector')
        client.province = request.form.get('province')
        client.is_final_consumer = is_final
        db.session.commit()
        flash('Cliente actualizado')
        return redirect(url_for('clients'))
    return render_template('cliente_form.html', client=client)

@app.post('/api/clients')
def api_create_client():
    data = request.get_json() or {}
    if not data.get('name'):
        return {'error': 'El nombre es obligatorio'}, 400
    is_final = data.get('type') == 'final'
    identifier = data.get('identifier') if not is_final else data.get('identifier') or None
    last_name = data.get('last_name') if is_final else None
    if not is_final and not identifier:
        return {'error': 'El RNC es obligatorio para empresas'}, 400
    if identifier:
        exists = company_query(Client).filter(Client.identifier == identifier).first()
        if exists:
            return {'error': 'Identifier already exists'}, 400
    email = data.get('email')
    if email:
        exists = company_query(Client).filter(Client.email == email).first()
        if exists:
            return {'error': 'Email already exists'}, 400
    client = Client(
        name=data.get('name'),
        last_name=last_name,
        identifier=identifier,
        phone=data.get('phone'),
        email=data.get('email'),
        street=data.get('street'),
        sector=data.get('sector'),
        province=data.get('province'),
        is_final_consumer=is_final,
        company_id=current_company_id()
    )
    db.session.add(client)
    db.session.commit()
    return {'id': client.id, 'name': client.name, 'identifier': client.identifier}


@app.get('/api/reference')
def api_reference():
    name = request.args.get('name', '')
    return {'reference': generate_reference(name)}

# Products CRUD
@app.route('/productos', methods=['GET', 'POST'])
def products():
    if request.method == 'POST':
        reference = request.form.get('reference') or generate_reference(request.form['name'])
        use_cost = bool(request.form.get('use_cost'))
        price = _to_float(request.form['price'])
        valid_cost, cost_price, warning_msg = _validate_product_cost_inputs(
            price,
            use_cost,
            request.form.get('cost_price'),
        )
        if not valid_cost:
            flash('No se pudo guardar el producto: costo inválido. Debe ser mayor que 0.')
            return redirect(url_for('products'))
        product = Product(
            code=request.form['code'],
            reference=reference,
            name=request.form['name'],
            unit=request.form['unit'],
            price=price,
            cost_price=cost_price,
            category=request.form.get('category'),
            has_itbis=bool(request.form.get('has_itbis')),
            company_id=current_company_id()
        )
        db.session.add(product)
        db.session.flush()
        _log_product_price_change(product, None, None)
        db.session.commit()
        flash('Producto agregado')
        if warning_msg:
            flash(warning_msg)
        notify('Producto agregado')
        return redirect(url_for('products'))
    cat = request.args.get('cat')
    query = company_query(Product)
    if cat:
        query = query.filter_by(category=cat)
    products = query.all()
    return render_template('productos.html', products=products, units=UNITS, categories=CATEGORIES, current_cat=cat)


@app.route('/productos/importar', methods=['GET', 'POST'])
@manager_only
def products_import():
    if request.method == 'POST':
        file = request.files['file']
        rows = file.stream.read().decode('utf-8').splitlines()
        reader = csv.DictReader(rows)
        for row in reader:
            code = row.get('code')
            if not code:
                continue
            prod = company_query(Product).filter_by(code=code).first()
            if not prod:
                prod = Product(code=code, company_id=current_company_id())
                db.session.add(prod)
            prod.name = row.get('name') or prod.name
            prod.unit = row.get('unit') or prod.unit
            prod.price = _to_float(row.get('price')) or prod.price
            cat = row.get('category')
            if cat in CATEGORIES:
                prod.category = cat
            prod.has_itbis = row.get('has_itbis', '').strip().lower() in ('1', 'true', 'si', 'sí', 'yes')
            if not prod.reference:
                prod.reference = generate_reference(prod.name)
        db.session.commit()
        flash('Productos importados')
        return redirect(url_for('products'))
    return render_template('productos_importar.html')


@app.route('/productos/export')
def export_products():
    """Export product catalog as CSV, optionally filtered by category."""
    cat = request.args.get('cat')
    query = company_query(Product)
    if cat:
        query = query.filter_by(category=cat)
    products = query.order_by(Product.name.asc()).all()

    mem = StringIO()
    writer = csv.writer(mem)
    writer.writerow(['code', 'reference', 'name', 'unit', 'price', 'cost_price', 'category', 'has_itbis'])
    for p in products:
        writer.writerow([
            p.code,
            p.reference or '',
            p.name,
            p.unit,
            p.price,
            p.cost_price if p.cost_price is not None else '',
            p.category or '',
            '1' if p.has_itbis else '0',
        ])
    mem.seek(0)
    return Response(
        mem.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=productos.csv'}
    )

@app.route('/productos/delete/<int:product_id>')
def delete_product(product_id):
    product = company_get(Product, product_id)
    db.session.delete(product)
    db.session.commit()
    flash('Producto eliminado')
    return redirect(url_for('products'))


@app.route('/inventario')
def inventory_report():
    wid = request.args.get('warehouse_id', type=int)
    q = request.args.get('q', '').strip()
    category = request.args.get('category', '')
    status = request.args.get('status', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)

    warehouses = company_query(Warehouse).order_by(Warehouse.name).all()
    stocks = []
    pagination = None
    movements = []
    if not wid and warehouses:
        wid = warehouses[0].id
    if wid:
        query = (
            company_query(ProductStock)
            .filter_by(warehouse_id=wid)
            .join(Product)
        )
        if q:
            like = f"%{q}%"
            query = query.filter(or_(Product.name.ilike(like), Product.code.ilike(like)))
        if category:
            query = query.filter(Product.category == category)
        if status == 'low':
            query = query.filter(ProductStock.stock > 0, ProductStock.stock <= ProductStock.min_stock)
        elif status == 'zero':
            query = query.filter(ProductStock.stock == 0)
        elif status == 'normal':
            query = query.filter(ProductStock.stock > ProductStock.min_stock)

        pagination = (
            query.order_by(Product.name)
            .paginate(page=page, per_page=per_page, error_out=False)
        )
        stocks = pagination.items
        movements = (
            company_query(InventoryMovement)
            .filter_by(warehouse_id=wid)
            .order_by(InventoryMovement.timestamp.desc())
            .limit(20)
            .all()
        )

    sales_total = (
        db.session.query(func.sum(Invoice.total))
        .filter_by(company_id=current_company_id(), warehouse_id=wid)
        .scalar()
        or 0
    )
    return render_template(
        'inventario.html',
        stocks=stocks,
        warehouses=warehouses,
        selected=wid,
        sales_total=sales_total,
        pagination=pagination,
        q=q,
        category=category,
        status=status,
        categories=CATEGORIES,
        per_page=per_page,
        movements=movements,
    )


@app.post('/inventario/<int:stock_id>/minimo')
def update_min_stock(stock_id):
    stock = company_get(ProductStock, stock_id)
    stock.min_stock = _to_int(request.form.get('min_stock'))
    db.session.commit()
    flash('Mínimo actualizado')
    return redirect(url_for('inventory_report', warehouse_id=stock.warehouse_id))


@app.route('/inventario/ajustar', methods=['GET', 'POST'])
def inventory_adjust():
    products = company_query(Product).order_by(Product.name).all()
    warehouses = company_query(Warehouse).order_by(Warehouse.name).all()
    if request.method == 'POST':
        pid = int(request.form['product_id'])
        wid = int(request.form['warehouse_id'])
        qty = _to_int(request.form['quantity'])
        mtype = request.form['movement_type']
        product = company_get(Product, pid)
        stock = company_query(ProductStock).filter_by(product_id=pid, warehouse_id=wid).first()
        if not stock:
            stock = ProductStock(product_id=pid, warehouse_id=wid, company_id=current_company_id())
            db.session.add(stock)
        # ensure numeric defaults
        if stock.stock is None:
            stock.stock = 0
        if product.stock is None:
            product.stock = 0
        if mtype == 'entrada':
            stock.stock += qty
            product.stock += qty
            mov_qty = qty
        elif mtype == 'salida':
            if stock.stock < qty:
                flash('Stock insuficiente')
                return redirect(url_for('inventory_adjust'))
            stock.stock -= qty
            product.stock -= qty
            mov_qty = qty
        else:  # ajuste
            mov_qty = abs(stock.stock - qty)
            product.stock += qty - stock.stock
            stock.stock = qty
        mov = InventoryMovement(
            product_id=product.id,
            quantity=mov_qty,
            movement_type=mtype,
            warehouse_id=wid,
            company_id=current_company_id(),
            executed_by=session.get('user_id'),
        )
        db.session.add(mov)
        db.session.commit()
        flash('Inventario actualizado')
        return redirect(url_for('inventory_report', warehouse_id=wid))
    return render_template('inventario_ajuste.html', products=products, warehouses=warehouses)


@app.route('/inventario/importar', methods=['GET', 'POST'])
def inventory_import():
    warehouses = company_query(Warehouse).order_by(Warehouse.name).all()
    if not warehouses:
        default = Warehouse(name='Principal', company_id=current_company_id())
        db.session.add(default)
        db.session.commit()
        warehouses = [default]
    if request.method == 'POST':
        wid = int(request.form['warehouse_id'])
        file = request.files.get('file')
        if not file or not file.filename.lower().endswith('.csv'):
            flash('Debe subir un archivo CSV válido')
            return render_template('inventario_importar.html', warehouses=warehouses)

        stream = StringIO(file.stream.read().decode('utf-8'))
        reader = csv.DictReader(stream)
        expected = {'code', 'stock', 'min_stock'}
        if not reader.fieldnames or not expected.issubset(set(reader.fieldnames)):
            flash('Cabeceras inválidas. Se requieren: code, stock, min_stock')
            return render_template('inventario_importar.html', warehouses=warehouses)

        errors = []
        valid_rows = []
        for idx, row in enumerate(reader, start=2):
            code = (row.get('code') or '').strip()
            if not code:
                errors.append((idx, 'Código faltante'))
                continue
            product = company_query(Product).filter_by(code=code).first()
            if not product:
                errors.append((idx, f'Producto {code} no encontrado'))
                continue
            try:
                stock_qty = int(row.get('stock'))
            except (TypeError, ValueError):
                errors.append((idx, f'Stock inválido para {code}'))
                continue
            min_val = row.get('min_stock')
            try:
                min_stock = int(min_val) if min_val not in (None, '') else None
            except ValueError:
                errors.append((idx, f'Min stock inválido para {code}'))
                continue
            valid_rows.append((product, stock_qty, min_stock))

        if errors:
            db.session.rollback()
            flash(f'Importación cancelada. {len(errors)} filas con errores.')
            return render_template('inventario_importar.html', warehouses=warehouses, errors=errors)

        for product, stock_qty, min_stock in valid_rows:
            product.stock = stock_qty
            ps = (
                company_query(ProductStock)
                .filter_by(product_id=product.id, warehouse_id=wid)
                .first()
            )
            if not ps:
                ps = ProductStock(product_id=product.id, warehouse_id=wid, company_id=current_company_id())
                db.session.add(ps)
            ps.stock = stock_qty
            if min_stock is not None:
                ps.min_stock = min_stock
                product.min_stock = min_stock
            mov = InventoryMovement(
                product_id=product.id,
                quantity=stock_qty,
                movement_type='entrada',
                reference_type='import',
                warehouse_id=wid,
                company_id=current_company_id(),
                executed_by=session.get('user_id'),
            )
            db.session.add(mov)

        db.session.commit()
        flash(f'Se importaron {len(valid_rows)} productos')
        return redirect(url_for('inventory_report', warehouse_id=wid))

    return render_template('inventario_importar.html', warehouses=warehouses)


@app.route('/inventario/transferir', methods=['GET', 'POST'])
def inventory_transfer():
    products = company_query(Product).order_by(Product.name).all()
    warehouses = company_query(Warehouse).order_by(Warehouse.name).all()
    if request.method == 'POST':
        pid = int(request.form['product_id'])
        origin = int(request.form['origin_id'])
        dest = int(request.form['dest_id'])
        qty = _to_int(request.form['quantity'])
        if origin == dest:
            flash('Seleccione almacenes distintos')
            return redirect(url_for('inventory_transfer'))
        o_stock = (
            company_query(ProductStock)
            .filter_by(product_id=pid, warehouse_id=origin)
            .first()
        )
        d_stock = (
            company_query(ProductStock)
            .filter_by(product_id=pid, warehouse_id=dest)
            .first()
        )
        if not o_stock or o_stock.stock < qty:
            flash('Stock insuficiente')
            return redirect(url_for('inventory_transfer'))
        if not d_stock:
            d_stock = ProductStock(product_id=pid, warehouse_id=dest, stock=0, company_id=current_company_id())
            db.session.add(d_stock)
        o_stock.stock -= qty
        d_stock.stock += qty
        mov_out = InventoryMovement(
            product_id=pid,
            quantity=qty,
            movement_type='salida',
            warehouse_id=origin,
            company_id=current_company_id(),
            reference_type='transfer',
            reference_id=dest,
            executed_by=session.get('user_id'),
        )
        mov_in = InventoryMovement(
            product_id=pid,
            quantity=qty,
            movement_type='entrada',
            warehouse_id=dest,
            company_id=current_company_id(),
            reference_type='transfer',
            reference_id=origin,
            executed_by=session.get('user_id'),
        )
        db.session.add_all([mov_out, mov_in])
        db.session.commit()
        flash('Transferencia realizada')
        return redirect(url_for('inventory_report', warehouse_id=dest))
    return render_template('inventario_transferir.html', products=products, warehouses=warehouses)


@app.route('/almacenes', methods=['GET', 'POST'])
@manager_only
def warehouses():
    if request.method == 'POST':
        name = request.form['name']
        address = request.form.get('address')
        w = Warehouse(name=name, address=address, company_id=current_company_id())
        db.session.add(w)
        db.session.commit()
        return redirect(url_for('warehouses'))
    ws = company_query(Warehouse).order_by(Warehouse.name).all()
    return render_template('almacenes.html', warehouses=ws)


@app.post('/almacenes/<int:w_id>/delete')
@manager_only
def delete_warehouse(w_id):
    w = company_get(Warehouse, w_id)
    db.session.delete(w)
    db.session.commit()
    return redirect(url_for('warehouses'))

@app.route('/productos/edit/<int:product_id>', methods=['GET', 'POST'])
def edit_product(product_id):
    product = company_get(Product, product_id)
    if request.method == 'POST':
        old_price = product.price
        old_cost_price = product.cost_price
        product.code = request.form['code']
        product.reference = request.form.get('reference') or generate_reference(request.form['name'])
        product.name = request.form['name']
        product.unit = request.form['unit']
        product.price = _to_float(request.form['price'])
        valid_cost, cost_price, warning_msg = _validate_product_cost_inputs(
            product.price,
            bool(request.form.get('use_cost')),
            request.form.get('cost_price'),
        )
        if not valid_cost:
            flash('No se pudo actualizar el producto: costo inválido. Debe ser mayor que 0.')
            return redirect(url_for('edit_product', product_id=product_id))
        product.cost_price = cost_price
        product.category = request.form.get('category')
        product.has_itbis = bool(request.form.get('has_itbis'))
        _log_product_price_change(product, old_price, old_cost_price)
        db.session.commit()
        flash('Producto actualizado')
        if warning_msg:
            flash(warning_msg)
        return redirect(url_for('products'))
    return render_template('producto_form.html', product=product, units=UNITS, categories=CATEGORIES)



@app.route('/productos/historial-precios')
def product_price_history():
    product_id = request.args.get('product_id', type=int)
    logs_q = (
        company_query(ProductPriceLog)
        .options(joinedload(ProductPriceLog.product), joinedload(ProductPriceLog.user))
        .order_by(ProductPriceLog.changed_at.desc())
    )
    if product_id:
        logs_q = logs_q.filter(ProductPriceLog.product_id == product_id)
    logs = logs_q.limit(200).all()
    products = company_query(Product).order_by(Product.name).all()
    return render_template('product_price_history.html', logs=logs, products=products, current_product_id=product_id)


# Quotations
@app.route('/cotizaciones')
def list_quotations():
    client_q = request.args.get('client')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    status = request.args.get('status')
    page = request.args.get('page', 1, type=int)

    now = dom_now()
    company_query(Quotation).filter(
        Quotation.status == 'vigente', Quotation.valid_until < now
    ).update({'status': 'vencida'}, synchronize_session=False)
    db.session.commit()

    query = company_query(Quotation).join(Client)
    if client_q:
        query = query.filter(
            (Client.name.contains(client_q)) | (Client.identifier.contains(client_q))
        )
    if date_from:
        df = datetime.strptime(date_from, '%Y-%m-%d')
        query = query.filter(Quotation.date >= df)
    if date_to:
        dt = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Quotation.date < dt)
    if status:
        query = query.filter(Quotation.status == status)

    quotations = query.order_by(Quotation.date.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    return render_template(
        'cotizaciones.html',
        quotations=quotations,
        client=client_q,
        date_from=date_from,
        date_to=date_to,
        status=status,
        now=now,
    )

@app.route('/cotizaciones/nueva', methods=['GET', 'POST'])
def new_quotation():
    if request.method == 'POST':
        print('Form data:', dict(request.form))
        client_id = request.form.get('client_id')
        if not client_id:
            flash('Debe seleccionar un cliente registrado')
            return redirect(url_for('new_quotation'))
        client = company_get(Client, client_id)
        wid = request.form.get('warehouse_id')
        if not wid:
            flash('Seleccione un almacén')
            return redirect(url_for('new_quotation'))
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('product_quantity[]')
        discounts = request.form.getlist('product_discount[]')
        items = build_items(product_ids, quantities, discounts)
        if not items:
            flash('Debe agregar al menos un producto')
            return redirect(url_for('new_quotation'))
        subtotal, itbis, total = calculate_totals(items)
        payment_method = request.form.get('payment_method')
        bank = request.form.get('bank') if payment_method == 'Transferencia' else None
        date = dom_now()
        validity_days = _quotation_validity_days(request.form.get('validity_period'))
        valid_until = date + timedelta(days=validity_days)
        quotation = Quotation(client_id=client.id, subtotal=subtotal, itbis=itbis, total=total,
                               seller=request.form.get('seller'), payment_method=payment_method,
                               bank=bank, note=request.form.get('note'),
                               warehouse_id=int(wid),
                               company_id=current_company_id(),
                               date=date, valid_until=valid_until)
        db.session.add(quotation)
        db.session.flush()
        for it in items:
            q_item = QuotationItem(quotation_id=quotation.id, **it)
            db.session.add(q_item)
        db.session.commit()
        flash('Cotización guardada')
        notify('Cotización guardada')
        log_audit('quotation_create', 'quotation', quotation.id, details=f'client={client.id};total={total:.2f}')

        company = get_company_info()
        quotation_pdf_bytes = _build_quotation_pdf_bytes(quotation, company)
        _archive_pdf_copy(
            'cotizacion',
            quotation.id,
            quotation_pdf_bytes,
            company_name=company.get('name'),
            company_id=current_company_id(),
        )

        is_first_quotation = company_query(Quotation).count() == 1
        if is_first_quotation:
            public_url = _public_doc_url(
                'cotizacion',
                quotation.id,
                company_name=company.get('name'),
                company_id=current_company_id(),
            )
            if public_url:
                return redirect(public_url)
        return redirect(url_for('list_quotations'))
    clients = company_query(Client).options(
        load_only(Client.id, Client.name, Client.identifier)
    ).all()
    products = company_query(Product).options(
        load_only(Product.id, Product.code, Product.name, Product.unit, Product.price)
    ).all()
    warehouses = company_query(Warehouse).order_by(Warehouse.name).all()
    sellers = company_query(User).options(load_only(User.id, User.first_name, User.last_name)).all()
    return render_template('cotizacion.html', clients=clients, products=products, warehouses=warehouses, sellers=sellers, validity_options=QUOTATION_VALIDITY_OPTIONS)

@app.route('/cotizaciones/editar/<int:quotation_id>', methods=['GET', 'POST'])
def edit_quotation(quotation_id):
    quotation = company_get(Quotation, quotation_id)
    if request.method == 'POST':
        client = quotation.client
        is_final = request.form.get('client_type') == 'final'
        identifier = request.form.get('client_identifier') if not is_final else request.form.get('client_identifier') or None
        if not is_final and not identifier:
            flash('El identificador es obligatorio para comprobante fiscal')
            return redirect(url_for('edit_quotation', quotation_id=quotation.id))
        client.name = request.form['client_name']
        client.last_name = request.form.get('client_last_name') if is_final else None
        client.identifier = identifier
        client.phone = request.form.get('client_phone')
        client.email = request.form.get('client_email')
        client.street = request.form.get('client_street')
        client.sector = request.form.get('client_sector')
        client.province = request.form.get('client_province')
        client.is_final_consumer = is_final
        quotation.items.clear()
        db.session.flush()
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('product_quantity[]')
        discounts = request.form.getlist('product_discount[]')
        items = build_items(product_ids, quantities, discounts)
        subtotal, itbis, total = calculate_totals(items)
        payment_method = request.form.get('payment_method')
        bank = request.form.get('bank') if payment_method == 'Transferencia' else None
        quotation.client_id = client.id
        quotation.subtotal = subtotal
        quotation.itbis = itbis
        quotation.total = total
        quotation.seller = request.form.get('seller')
        quotation.payment_method = payment_method
        quotation.bank = bank
        quotation.note = request.form.get('note')
        validity_days = _quotation_validity_days(request.form.get('validity_period'))
        quotation.valid_until = (quotation.date or dom_now()) + timedelta(days=validity_days)
        for it in items:
            quotation.items.append(QuotationItem(**it))
        db.session.commit()
        flash('Cotización actualizada')
        log_audit('quotation_update', 'quotation', quotation.id, details=f'total={total:.2f}')
        return redirect(url_for('list_quotations'))
    products = company_query(Product).options(
        load_only(Product.id, Product.code, Product.name, Product.unit, Product.price)
    ).all()
    product_map = {p.name: p.id for p in products}
    items = []
    for it in quotation.items:
        base = it.unit_price * it.quantity
        percent = (it.discount / base * 100) if base else 0
        items.append({
            'product_id': product_map.get(it.product_name, ''),
            'quantity': it.quantity,
            'discount': percent,
            'unit': it.unit,
            'price': it.unit_price,
        })
    sellers = company_query(User).options(load_only(User.id, User.first_name, User.last_name)).all()
    return render_template(
        'cotizacion_edit.html',
        quotation=quotation,
        products=products,
        items=items,
        sellers=sellers,
        validity_options=QUOTATION_VALIDITY_OPTIONS,
    )


@app.route('/ajustes')
@manager_only
def settings():
    return redirect(url_for('settings_company'))


@app.route('/ajustes/empresa', methods=['GET', 'POST'])
@manager_only
def settings_company():
    company = db.session.get(CompanyInfo, current_company_id())
    if not company:
        flash('Seleccione una empresa')
        return redirect(url_for('admin_companies'))
    if request.method == 'POST':
        role = session.get('role')
        if role != 'manager':
            company.name = request.form.get('name', company.name)
            company.street = request.form.get('street', company.street)
            company.sector = request.form.get('sector', company.sector)
            company.province = request.form.get('province', company.province)
            company.phone = request.form.get('phone', company.phone)
            company.rnc = request.form.get('rnc', company.rnc)
            company.website = request.form.get('website') or None
        if request.form.get('remove_logo'):
            if company.logo:
                try:
                    os.remove(os.path.join(app.static_folder, company.logo))
                except FileNotFoundError:
                    pass
            company.logo = None
        else:
            file = request.files.get('logo')
            if file and file.filename:
                filename = secure_filename(file.filename)
                ext = os.path.splitext(filename)[1].lower()
                if ext not in {'.png', '.jpg', '.jpeg'}:
                    flash('Formato de logo inválido')
                    return redirect(url_for('settings_company'))
                file.seek(0, os.SEEK_END)
                size = file.tell()
                file.seek(0)
                if size > 1 * 1024 * 1024:
                    flash('Logo demasiado grande (máximo 1MB)')
                    return redirect(url_for('settings_company'))
                upload_dir = os.path.join(app.static_folder, 'uploads')
                os.makedirs(upload_dir, exist_ok=True)
                path = os.path.join(upload_dir, filename)
                file.save(path)
                company.logo = f'uploads/{filename}'
        old_final = company.ncf_final
        old_fiscal = company.ncf_fiscal
        new_final = _to_int(request.form.get('ncf_final'))
        new_fiscal = _to_int(request.form.get('ncf_fiscal'))
        if new_final is not None and new_final < old_final:
            flash('NCF Consumidor Final no puede ser menor que el actual')
            return redirect(url_for('settings_company'))
        if new_fiscal is not None and new_fiscal < old_fiscal:
            flash('NCF Comprobante Fiscal no puede ser menor que el actual')
            return redirect(url_for('settings_company'))
        if new_final is not None:
            company.ncf_final = new_final
        if new_fiscal is not None:
            company.ncf_fiscal = new_fiscal
        if old_final != company.ncf_final or old_fiscal != company.ncf_fiscal:
            log = NcfLog(
                company_id=company.id,
                old_final=old_final,
                old_fiscal=old_fiscal,
                new_final=company.ncf_final,
                new_fiscal=company.ncf_fiscal,
                changed_by=session.get('user_id'),
            )
            db.session.add(log)
        db.session.commit()
        flash('Ajustes guardados')
        return redirect(url_for('settings_company'))
    owner_user = (
        User.query.filter_by(company_id=company.id, role='admin')
        .order_by(User.id.asc())
        .first()
    )
    if not owner_user:
        owner_user = User.query.filter_by(company_id=company.id).order_by(User.id.asc()).first()
    owner_email = owner_user.email if owner_user else ''
    return render_template('ajustes_empresa.html', company=company, owner_email=owner_email)


@app.route('/ajustes/usuarios/agregar', methods=['GET', 'POST'])
@manager_only
def settings_add_user():
    company = db.session.get(CompanyInfo, current_company_id())
    if request.method == 'POST':
        if session.get('role') == 'manager':
            count = User.query.filter_by(company_id=company.id, role='company').count()
            if count >= 2:
                flash('Los managers solo pueden crear 2 usuarios')
                return redirect(url_for('settings_add_user'))
        username = (request.form.get('username') or '').strip().lower()
        password = request.form.get('password') or ''
        if not username:
            flash('Debe ingresar un usuario válido')
            return redirect(url_for('settings_add_user'))
        if len(password) < 6 and not app.config.get('TESTING', False):
            flash('La contraseña debe tener mínimo 6 caracteres')
            return redirect(url_for('settings_add_user'))
        if User.query.filter(func.lower(User.username) == username).first():
            flash('Ese usuario ya existe')
            return redirect(url_for('settings_add_user'))
        user = User(
            username=username,
            first_name=request.form['first_name'],
            last_name=request.form['last_name'],
            role='company',
            company_id=company.id,
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash('Usuario creado')
        return redirect(url_for('settings_manage_users'))
    return render_template('ajustes_usuario_form.html')


@app.route('/ajustes/usuarios', methods=['GET', 'POST'])
@manager_only
def settings_manage_users():
    company_id = current_company_id()
    if request.method == 'POST':
        uid = int(request.form['user_id'])
        user = company_get(User, uid)
        action = request.form.get('action')
        if action == 'delete':
            db.session.delete(user)
            db.session.commit()
            flash('Usuario eliminado')
            return redirect(url_for('settings_manage_users'))
        user.first_name = request.form.get('first_name', user.first_name)
        user.last_name = request.form.get('last_name', user.last_name)
        new_username = (request.form.get('username', user.username) or '').strip().lower()
        if not new_username:
            flash('Usuario inválido')
            return redirect(url_for('settings_manage_users'))
        conflict = User.query.filter(func.lower(User.username) == new_username, User.id != user.id).first()
        if conflict:
            flash('Ese usuario ya existe')
            return redirect(url_for('settings_manage_users'))
        user.username = new_username
        new_role = request.form.get('role', user.role)
        if new_role in ('company', 'manager'):
            user.role = new_role
        db.session.commit()
        flash('Usuario actualizado')
        return redirect(url_for('settings_manage_users'))
    users = (
        User.query.filter_by(company_id=company_id)
        .filter(User.id != session.get('user_id'))
        .all()
    )
    return render_template('ajustes_usuarios.html', users=users)

@app.route('/cotizaciones/<int:quotation_id>/pdf')
def quotation_pdf(quotation_id):
    quotation = company_get(Quotation, quotation_id)
    company = get_company_info()
    filename = f'cotizacion_{quotation_id}.pdf'
    app.logger.info("Generating quotation PDF %s", quotation_id)
    try:
        pdf_data = _build_quotation_pdf_bytes(quotation, company)
        return _archive_and_send_pdf(
            doc_type='cotizacion',
            doc_number=quotation.id,
            pdf_data=pdf_data,
            download_name=filename,
            company_name=company.get('name'),
        )
    except Exception as exc:
        app.logger.exception('Quotation PDF generation failed id=%s: %s', quotation_id, exc)
        archived = _archived_pdf_path(
            'cotizacion',
            quotation.id,
            company_name=company.get('name'),
            company_id=current_company_id(),
        )
        if archived.exists():
            try:
                return send_file(str(archived), as_attachment=True, download_name=filename, mimetype='application/pdf')
            except TypeError:
                return send_file(str(archived), as_attachment=True, attachment_filename=filename, mimetype='application/pdf')
        raise


@app.route('/cotizaciones/<int:quotation_id>/enviar', methods=['POST'])
def send_quotation_email(quotation_id):
    quotation = company_get(Quotation, quotation_id)
    client = quotation.client
    if not client.email:
        flash('El cliente no tiene correo registrado')
        return redirect(url_for('list_quotations'))
    company = get_company_info()
    filename = f'cotizacion_{quotation_id}.pdf'
    pdf_data = generate_pdf_bytes('Cotización', company, client, quotation.items,
                                  quotation.subtotal, quotation.itbis, quotation.total,
                                  seller=quotation.seller, payment_method=quotation.payment_method,
                                  bank=quotation.bank, doc_number=quotation.id, note=quotation.note,
                                  date=quotation.date, valid_until=quotation.valid_until,
                                  footer=("Condiciones: Esta cotización es válida por 30 días a partir de la fecha de emisión. "
                                          "Los precios están sujetos a cambios sin previo aviso. "
                                          "El ITBIS ha sido calculado conforme a la ley vigente."))
    html = render_template('emails/quotation.html', client=client, company=company, quotation=quotation)
    send_email(client.email, 'Cotización', html, attachments=[(filename, pdf_data)])
    flash(f'Cotización enviada con éxito a {client.email}')
    return redirect(url_for('list_quotations'))

@app.route('/cotizaciones/<int:quotation_id>/convertir', methods=['GET', 'POST'])
def quotation_to_order(quotation_id):
    quotation = company_get(Quotation, quotation_id)
    warehouses = company_query(Warehouse).all()
    if request.method == 'GET':
        return render_template('quotation_convert.html', quotation=quotation, warehouses=warehouses)
    wid = request.form.get('warehouse_id', quotation.warehouse_id)
    if not wid:
        flash('Seleccione un almacén')
        return redirect(url_for('quotation_to_order', quotation_id=quotation_id))
    wid = int(wid)
    quotation.warehouse_id = wid
    customer_po = request.form.get('customer_po') or None
    if dom_now() > quotation.valid_until:
        flash('La cotización ha expirado')
        return redirect(url_for('list_quotations'))
    for item in quotation.items:
        product = company_query(Product).filter_by(code=item.code).first()
        stock = (
            company_query(ProductStock)
            .filter_by(product_id=product.id, warehouse_id=wid)
            .first()
        )
        if not stock or stock.stock < item.quantity:
            flash('Stock insuficiente para ' + item.product_name)
            return redirect(url_for('list_quotations'))
    order = Order(
        client_id=quotation.client_id,
        quotation_id=quotation.id,
        subtotal=quotation.subtotal,
        itbis=quotation.itbis,
        total=quotation.total,
        seller=quotation.seller,
        payment_method=quotation.payment_method,
        bank=quotation.bank,
        note=quotation.note,
        customer_po=customer_po,
        warehouse_id=wid,
        company_id=current_company_id(),
    )
    db.session.add(order)
    quotation.status = 'convertida'
    db.session.flush()
    for item in quotation.items:
        o_item = OrderItem(
            order_id=order.id,
            code=item.code,
            reference=item.reference,
            product_name=item.product_name,
            unit=item.unit,
            unit_price=item.unit_price,
            quantity=item.quantity,
            discount=item.discount,
            category=item.category,
            has_itbis=item.has_itbis,
            company_id=current_company_id(),
        )
        db.session.add(o_item)
        product = company_query(Product).filter_by(code=item.code).first()
        if product:
            ps = (
                company_query(ProductStock)
                .filter_by(product_id=product.id, warehouse_id=wid)
                .first()
            )
            if ps:
                ps.stock -= item.quantity
            product.stock -= item.quantity
            mov = InventoryMovement(
                product_id=product.id,
                quantity=item.quantity,
                movement_type='salida',
                reference_type='Order',
                reference_id=order.id,
                warehouse_id=wid,
                company_id=current_company_id(),
                executed_by=session.get('user_id'),
            )
            db.session.add(mov)
    db.session.commit()
    company = get_company_info()
    order_pdf_bytes = _build_order_pdf_bytes(order, company)
    _archive_pdf_copy(
        'pedido',
        order.id,
        order_pdf_bytes,
        company_name=company.get('name'),
        company_id=current_company_id(),
    )
    flash('Pedido creado')
    notify('Pedido creado')
    return redirect(url_for('list_orders'))

# Orders
@app.route('/pedidos')
def list_orders():
    q = request.args.get('q')
    query = company_query(Order).join(Client)
    if q:
        query = query.filter((Client.name.contains(q)) | (Client.identifier.contains(q)))
    orders = query.order_by(Order.date.desc()).all()
    return render_template('pedido.html', orders=orders, q=q)

@app.route('/pedidos/<int:order_id>/facturar')
def order_to_invoice(order_id):
    order = company_get(Order, order_id)
    company = db.session.get(CompanyInfo, current_company_id())
    if order.client.is_final_consumer:
        prefix, counter = "B02", "ncf_final"
    else:
        prefix, counter = "B01", "ncf_fiscal"

    # ensure NCF is unique; advance company counter until unused
    while True:
        seq = getattr(company, counter)
        ncf = f"{prefix}{seq:08d}"
        if not Invoice.query.filter_by(ncf=ncf).first():
            setattr(company, counter, seq + 1)
            break
        setattr(company, counter, seq + 1)
    invoice = Invoice(
        client_id=order.client_id,
        order_id=order.id,
        subtotal=order.subtotal,
        itbis=order.itbis,
        total=order.total,
        ncf=ncf,
        seller=order.seller,
        payment_method=order.payment_method,
        bank=order.bank,
        note=order.note,
        invoice_type=(
            'Consumidor Final' if order.client.is_final_consumer else 'Crédito Fiscal'
        ),
        status='Pendiente',
        warehouse_id=order.warehouse_id,
        company_id=current_company_id(),
    )
    db.session.add(invoice)
    db.session.flush()
    for item in order.items:
        i_item = InvoiceItem(
            invoice_id=invoice.id,
            code=item.code,
            reference=item.reference,
            product_name=item.product_name,
            unit=item.unit,
            unit_price=item.unit_price,
            quantity=item.quantity,
            discount=item.discount,
            category=item.category,
            has_itbis=item.has_itbis,
            company_id=current_company_id(),
        )
        db.session.add(i_item)
    order.status = 'Entregado'
    db.session.commit()
    company_info = get_company_info()
    invoice_pdf_bytes = _build_invoice_pdf_bytes(invoice, company_info)
    _archive_pdf_copy(
        'factura',
        invoice.id,
        invoice_pdf_bytes,
        company_name=company_info.get('name'),
        company_id=current_company_id(),
    )
    flash('Factura generada')
    notify('Factura generada')
    log_audit('invoice_create', 'invoice', invoice.id, details=f'from_order={order.id};total={invoice.total:.2f}')
    return redirect(url_for('list_invoices'))

@app.route('/pedidos/<int:order_id>/pdf')
def order_pdf(order_id):
    order = company_get(Order, order_id)
    company = get_company_info()
    filename = f'pedido_{order_id}.pdf'
    app.logger.info("Generating order PDF %s", order_id)
    pdf_data = _build_order_pdf_bytes(order, company)
    return _archive_and_send_pdf(
        doc_type='pedido',
        doc_number=order.id,
        pdf_data=pdf_data,
        download_name=filename,
        company_name=company.get('name'),
    )

# Invoices
@app.route('/facturas')
def list_invoices():
    q = request.args.get('q')
    query = company_query(Invoice).join(Client)
    if q:
        query = query.filter((Client.name.contains(q)) | (Client.identifier.contains(q)))
    invoices = query.order_by(Invoice.date.desc()).all()
    return render_template('factura.html', invoices=invoices, q=q)


@app.route('/facturas/<int:invoice_id>/pagar', methods=['POST'])
def pay_invoice(invoice_id):
    invoice = company_get(Invoice, invoice_id)
    invoice.status = 'Pagada'
    db.session.commit()
    flash('Factura marcada como pagada')
    log_audit('invoice_paid', 'invoice', invoice.id, details=f'total={invoice.total:.2f}')
    return redirect(url_for('list_invoices'))


@app.route('/notificaciones')
def notifications_view():
    notifs = company_query(Notification).order_by(Notification.created_at.desc()).all()
    return render_template('notifications.html', notifications=notifs)


@app.post('/notificaciones/<int:nid>/leer')
def notifications_read(nid):
    notif = company_get(Notification, nid)
    notif.is_read = True
    db.session.commit()
    return redirect(url_for('notifications_view'))

@app.route('/facturas/<int:invoice_id>/pdf')
def invoice_pdf(invoice_id):
    invoice = company_get(Invoice, invoice_id)
    company = get_company_info()
    filename = f'factura_{invoice_id}.pdf'
    app.logger.info("Generating invoice PDF %s", invoice_id)
    pdf_data = _build_invoice_pdf_bytes(invoice, company)
    return _archive_and_send_pdf(
        doc_type='factura',
        doc_number=invoice.id,
        pdf_data=pdf_data,
        download_name=filename,
        company_name=company.get('name'),
    )

@app.route('/pdfs/<path:filename>')
def serve_pdf(filename):
    if '..' in filename or filename.startswith('/'):
        return ('Not Found', 404)
    base = _company_pdf_dir().parent.resolve()
    file_path = (base / filename).resolve()
    if not str(file_path).startswith(str(base.resolve())) or not file_path.exists():
        return ('Not Found', 404)
    return send_file(str(file_path), as_attachment=True)

def _filtered_invoice_query(fecha_inicio, fecha_fin, estado, categoria):
    """Return an invoice query filtered by the provided parameters."""
    q = company_query(Invoice)
    if fecha_inicio:
        q = q.filter(Invoice.date >= fecha_inicio)
    if fecha_fin:
        q = q.filter(Invoice.date <= fecha_fin)
    if estado:
        q = q.filter(Invoice.status == estado)
    if categoria:
        q = q.join(Invoice.items).filter(InvoiceItem.category == categoria)
    return q


@app.route('/reportes')
def reportes():
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    estado = request.args.get('estado')
    categoria = request.args.get('categoria')
    page = request.args.get('page', 1, type=int)

    start, end, estado, categoria = _parse_report_params(fecha_inicio, fecha_fin, estado, categoria)
    q = _filtered_invoice_query(start, end, estado, categoria)

    pagination = (
        q.options(
            joinedload(Invoice.client),
            load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
        )
        .order_by(Invoice.date.desc())
        .paginate(page=page, per_page=10, error_out=False)
    )
    invoices = pagination.items

    total_sales, unique_clients, invoice_count = (
        q.with_entities(
            func.coalesce(func.sum(Invoice.total), 0),
            func.count(func.distinct(Invoice.client_id)),
            func.count(Invoice.id),
        ).first()
    )

    item_query = company_query(InvoiceItem).join(Invoice)
    if start:
        item_query = item_query.filter(Invoice.date >= start)
    if end:
        item_query = item_query.filter(Invoice.date <= end)
    if estado:
        item_query = item_query.filter(Invoice.status == estado)
    if categoria:
        item_query = item_query.filter(InvoiceItem.category == categoria)

    sales_by_category = (
        item_query.with_entities(
            InvoiceItem.category,
            func.count(InvoiceItem.id),
            func.avg((InvoiceItem.unit_price * InvoiceItem.quantity) - InvoiceItem.discount),
            func.sum((InvoiceItem.unit_price * InvoiceItem.quantity) - InvoiceItem.discount),
        ).group_by(InvoiceItem.category).all()
    )

    sales_over_time = (
        q.with_entities(func.date(Invoice.date), func.sum(Invoice.total), func.count(Invoice.id))
        .group_by(func.date(Invoice.date))
        .order_by(func.date(Invoice.date))
        .all()
    )

    # retention
    client_counts = (
        q.with_entities(Invoice.client_id, func.count(Invoice.id))
        .group_by(Invoice.client_id)
        .all()
    )
    retained = len([1 for _cid, cnt in client_counts if cnt > 1])
    retention = (retained / len(client_counts)) * 100 if client_counts else 0

    # top categories last year
    last_year_start = datetime.utcnow().replace(year=datetime.utcnow().year - 1, month=1, day=1)
    top_cats = (
        company_query(InvoiceItem)
        .join(Invoice)
        .with_entities(InvoiceItem.category, func.sum((InvoiceItem.unit_price * InvoiceItem.quantity) - InvoiceItem.discount))
        .filter(Invoice.date >= last_year_start)
        .group_by(InvoiceItem.category)
        .order_by(func.sum((InvoiceItem.unit_price * InvoiceItem.quantity) - InvoiceItem.discount).desc())
        .limit(5)
        .all()
    )

    # monthly and yearly avg ticket
    today = datetime.utcnow()
    month_total, month_clients = (
        q.filter(
            extract('year', Invoice.date) == today.year,
            extract('month', Invoice.date) == today.month,
        )
        .with_entities(
            func.coalesce(func.sum(Invoice.total), 0),
            func.count(func.distinct(Invoice.client_id)),
        )
        .first()
    )
    avg_ticket_month = month_total / month_clients if month_clients else 0
    year_total, year_clients = (
        q.filter(extract('year', Invoice.date) == today.year)
        .with_entities(
            func.coalesce(func.sum(Invoice.total), 0),
            func.count(func.distinct(Invoice.client_id)),
        )
        .first()
    )
    avg_ticket_year = year_total / year_clients if year_clients else 0

    itbis_accumulated, net_sales = (
        q.with_entities(
            func.coalesce(func.sum(Invoice.itbis), 0),
            func.coalesce(func.sum(Invoice.subtotal), 0),
        ).first()
    )

    profit_query = company_query(InvoiceItem).join(Invoice)
    if start:
        profit_query = profit_query.filter(Invoice.date >= start)
    if end:
        profit_query = profit_query.filter(Invoice.date <= end)
    if estado:
        profit_query = profit_query.filter(Invoice.status == estado)
    if categoria:
        profit_query = profit_query.filter(InvoiceItem.category == categoria)
    profit_base = profit_query.outerjoin(
        Product,
        (Product.company_id == InvoiceItem.company_id) & (Product.code == InvoiceItem.code),
    )
    estimated_profit_with_cost = (
        profit_base
        .filter(Product.cost_price.isnot(None))
        .with_entities(
            func.coalesce(
                func.sum(
                    ((InvoiceItem.unit_price - Product.cost_price) * InvoiceItem.quantity)
                    - InvoiceItem.discount
                ),
                0,
            )
        )
        .scalar()
    )
    revenue_without_cost_data = (
        profit_base
        .filter(Product.cost_price.is_(None))
        .with_entities(
            func.coalesce(
                func.sum((InvoiceItem.unit_price * InvoiceItem.quantity) - InvoiceItem.discount),
                0,
            )
        )
        .scalar()
    )
    estimated_profit = estimated_profit_with_cost

    kpi_changes = {
        'net_sales': None,
        'itbis_accumulated': None,
        'estimated_profit_with_cost': None,
    }
    if start and end:
        period_days = (end.date() - start.date()).days + 1
        prev_end = start - timedelta(days=1)
        prev_start = prev_end - timedelta(days=period_days - 1)

        q_prev = _filtered_invoice_query(prev_start, prev_end, estado, categoria)
        prev_itbis, prev_net_sales = (
            q_prev.with_entities(
                func.coalesce(func.sum(Invoice.itbis), 0),
                func.coalesce(func.sum(Invoice.subtotal), 0),
            ).first()
        )

        prev_profit_query = company_query(InvoiceItem).join(Invoice)
        prev_profit_query = prev_profit_query.filter(Invoice.date >= prev_start, Invoice.date <= prev_end)
        if estado:
            prev_profit_query = prev_profit_query.filter(Invoice.status == estado)
        if categoria:
            prev_profit_query = prev_profit_query.filter(InvoiceItem.category == categoria)
        prev_profit_with_cost = (
            prev_profit_query
            .outerjoin(
                Product,
                (Product.company_id == InvoiceItem.company_id) & (Product.code == InvoiceItem.code),
            )
            .filter(Product.cost_price.isnot(None))
            .with_entities(
                func.coalesce(
                    func.sum(
                        ((InvoiceItem.unit_price - Product.cost_price) * InvoiceItem.quantity)
                        - InvoiceItem.discount
                    ),
                    0,
                )
            )
            .scalar()
        )

        kpi_changes = {
            'net_sales': _pct_change(net_sales, prev_net_sales),
            'itbis_accumulated': _pct_change(itbis_accumulated, prev_itbis),
            'estimated_profit_with_cost': _pct_change(estimated_profit_with_cost, prev_profit_with_cost),
        }

    # trend last 24 months
    trend_query = (
        q.with_entities(
            extract('year', Invoice.date).label('y'),
            extract('month', Invoice.date).label('m'),
            func.sum(Invoice.total),
        )
        .filter(Invoice.date >= datetime(today.year - 2, today.month, 1))
        .group_by('y', 'm')
        .order_by('y', 'm')
    )
    # ensure a list of dicts is always provided even if the query returns no rows
    trend_rows = trend_query.all() if trend_query is not None else []
    trend_24 = [
        {'month': f"{int(y):04d}-{int(m):02d}", 'total': tot or 0}
        for y, m, tot in trend_rows
    ]

    status_totals = {s: 0 for s in INVOICE_STATUSES}
    status_counts = {s: 0 for s in INVOICE_STATUSES}
    for st, amount, cnt in (
        q.with_entities(Invoice.status, func.sum(Invoice.total), func.count(Invoice.id))
        .group_by(Invoice.status)
    ):
        if st in status_totals:
            status_totals[st] = amount or 0
            status_counts[st] = cnt or 0

    payment_totals = {'Efectivo': 0, 'Transferencia': 0}
    payment_counts = {'Efectivo': 0, 'Transferencia': 0}
    for pm, amount, cnt in (
        q.with_entities(Invoice.payment_method, func.sum(Invoice.total), func.count(Invoice.id))
        .group_by(Invoice.payment_method)
    ):
        if pm in payment_totals:
            payment_totals[pm] = amount or 0
            payment_counts[pm] = cnt or 0

    current_year = datetime.utcnow().year
    monthly_totals = (
        q.with_entities(
            extract('year', Invoice.date).label('y'),
            extract('month', Invoice.date).label('m'),
            func.sum(Invoice.total),
        )
        .filter(Invoice.date >= datetime(current_year - 1, 1, 1))
        .group_by('y', 'm')
        .all()
    )
    year_current = [0] * 12
    year_prev = [0] * 12
    for y, m, total in monthly_totals:
        if int(y) == current_year:
            year_current[int(m) - 1] = total or 0
        else:
            year_prev[int(m) - 1] = total or 0

    avg_ticket = total_sales / unique_clients if unique_clients else 0

    top_clients = (
        q.join(Client)
        .with_entities(Client.name, func.sum(Invoice.total))
        .group_by(Client.id)
        .order_by(func.sum(Invoice.total).desc())
        .limit(5)
        .all()
    )

    stats = {
        'total_sales': total_sales,
        'unique_clients': unique_clients,
        'invoices': invoice_count,
        'pending': status_totals.get('Pendiente', 0),
        'paid': status_totals.get('Pagada', 0),
        'cash': payment_totals.get('Efectivo', 0),
        'transfer': payment_totals.get('Transferencia', 0),
        'avg_ticket': avg_ticket,
        'avg_ticket_month': avg_ticket_month,
        'avg_ticket_year': avg_ticket_year,
        'retention': retention,
        'itbis_accumulated': itbis_accumulated,
        'net_sales': net_sales,
        'estimated_profit': estimated_profit,
        'estimated_profit_with_cost': estimated_profit_with_cost,
        'revenue_without_cost_data': revenue_without_cost_data,
    }

    cat_labels = [c or 'Sin categoría' for c, *_ in sales_by_category]
    cat_totals = [s or 0 for *_1, _2, s in sales_by_category]
    cat_counts = [qtd for _cat, qtd, *_ in sales_by_category]
    date_labels = [d if isinstance(d, str) else d.strftime('%Y-%m-%d') for d, *_ in sales_over_time]
    date_totals = [t or 0 for _, t, _ in sales_over_time]
    date_counts = [cnt for *_1, cnt in sales_over_time]
    status_labels = list(status_counts.keys())
    status_values = list(status_counts.values())
    method_labels = list(payment_counts.keys())
    method_values = list(payment_counts.values())
    months = [
        'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
        'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic'
    ]

    filters = {
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'estado': estado or '',
        'categoria': categoria or '',
    }

    if request.args.get('ajax') == '1':
        return jsonify(
            {
                'stats': stats,
                'top_clients': [{'name': n, 'total': t} for n, t in top_clients],
                'cat_labels': cat_labels,
                'cat_totals': cat_totals,
                'cat_counts': cat_counts,
                'date_labels': date_labels,
                'date_totals': date_totals,
                'date_counts': date_counts,
                'status_labels': status_labels,
                'status_values': status_values,
                'method_labels': method_labels,
                'method_values': method_values,
                'months': months,
                'year_current': year_current,
                'year_prev': year_prev,
                'top_categories_year': [{'category': c or 'Sin categoría', 'total': t or 0} for c, t in top_cats],
                'trend_24': trend_24,
                'invoices': [
                    {
                        'client': i.client.name if i.client else '',
                        'date': i.date.strftime('%Y-%m-%d'),
                        'estado': i.status or '',
                        'total': i.total,
                    }
                    for i in invoices
                ],
                'pagination': {'page': pagination.page, 'pages': pagination.pages},
                'kpi_changes': kpi_changes,
            }
        )

    return render_template(
        'reportes.html',
        invoices=invoices,
        pagination=pagination,
        sales_by_category=sales_by_category,
        stats=stats,
        top_clients=top_clients,
        top_categories_year=top_cats,
        trend_24=trend_24,
        cat_labels=cat_labels,
        cat_totals=cat_totals,
        cat_counts=cat_counts,
        date_labels=date_labels,
        date_totals=date_totals,
        date_counts=date_counts,
        status_labels=status_labels,
        status_values=status_values,
        method_labels=method_labels,
        method_values=method_values,
        kpi_changes=kpi_changes,
        months=months,
        year_current=year_current,
        year_prev=year_prev,
        filters=filters,
        categories=CATEGORIES,
        statuses=INVOICE_STATUSES,
    )


@app.get('/reportes/estado-cuentas')
def account_statement_clients():
    clients = company_query(Client).order_by(Client.name).all()
    return render_template('estado_cuentas.html', clients=clients)


def _invoice_balance(inv):
    return inv.total - sum(p.amount for p in inv.payments)


@app.get('/reportes/estado-cuentas/<int:client_id>')
def account_statement_detail(client_id):
    client = company_get(Client, client_id)
    invoices = (
        company_query(Invoice)
        .filter_by(client_id=client.id)
        .options(joinedload(Invoice.order), joinedload(Invoice.payments))
        .all()
    )
    rows = []
    totals = 0
    aging = {'0-30': 0, '31-60': 0, '61-90': 0, '91-120': 0, '121+': 0}
    now = dom_now()
    for inv in invoices:
        balance = _invoice_balance(inv)
        if balance <= 0:
            continue
        due = inv.date + timedelta(days=30)
        rows.append({
            'document': inv.ncf or f'FAC-{inv.id}',
            'order': inv.order.customer_po if inv.order and inv.order.customer_po else inv.order_id,
            'date': inv.date.strftime('%d/%m/%Y'),
            'due': due.strftime('%d/%m/%Y'),
            'info': inv.note or '',
            'amount': inv.total,
            'balance': balance,
        })
        totals += balance
        age = (now - inv.date).days
        if age <= 30:
            aging['0-30'] += balance
        elif age <= 60:
            aging['31-60'] += balance
        elif age <= 90:
            aging['61-90'] += balance
        elif age <= 120:
            aging['91-120'] += balance
        else:
            aging['121+'] += balance
    overdue = sum(r['balance'] for r in rows if datetime.strptime(r['due'], '%d/%m/%Y') < now)
    overdue_pct = (overdue / totals * 100) if totals else 0
    if request.args.get('pdf') == '1':
        company = {
            'name': g.company.name,
            'street': g.company.street,
            'phone': g.company.phone,
            'rnc': g.company.rnc,
            'logo': g.company.logo,
        }
        client_dict = {
            'name': client.name,
            'identifier': client.identifier,
            'street': client.street,
            'sector': client.sector,
            'province': client.province,
            'phone': client.phone,
            'email': client.email,
        }
        pdf_data = generate_account_statement_pdf_bytes(company, client_dict, rows, totals, aging, overdue_pct)
        return _archive_and_send_pdf(
            doc_type='estado_cuenta',
            doc_number=client.id,
            pdf_data=pdf_data,
            download_name=f'estado_cuenta_{client.id}.pdf',
            company_name=company.get('name'),
        )
    return render_template('estado_cuenta_detalle.html', client=client, rows=rows, total=totals, aging=aging, overdue_pct=overdue_pct)


@app.route('/reportes/export')
def export_reportes():
    role = session.get('role')
    formato = request.args.get('formato', 'csv')
    tipo = request.args.get('tipo', 'detalle')
    if role == 'contabilidad':
        if formato not in {'csv', 'xlsx'} or tipo != 'resumen':
            log_export(session.get('full_name') or session.get('username'), formato, tipo, {}, 'fail', 'permiso')
            return '', 403
    elif role not in ('admin', 'manager'):
        log_export(session.get('full_name') or session.get('username'), formato, tipo, {}, 'fail', 'permiso')
        return '', 403

    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    estado = request.args.get('estado')
    categoria = request.args.get('categoria')

    start, end, estado, categoria = _parse_report_params(fecha_inicio, fecha_fin, estado, categoria)
    q = _filtered_invoice_query(start, end, estado, categoria)
    count = q.count()
    filtros = {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin, 'estado': estado, 'categoria': categoria}
    user = session.get('full_name') or session.get('username')

    max_rows = current_app.config.get('MAX_EXPORT_ROWS', MAX_EXPORT_ROWS)
    if count > max_rows and request.args.get('async') != '1':
        log_export(user, formato, tipo, filtros, 'fail', 'too_many_rows')
        return jsonify({'error': 'too many rows', 'suggest': 'async'}), 400

    if count > max_rows and request.args.get('async') == '1':
        entry_id = log_export(user, formato, tipo, filtros, 'queued')
        enqueue_export(
            _export_job,
            current_company_id(),
            user,
            start,
            end,
            estado,
            categoria,
            formato,
            tipo,
            entry_id,
        )
        flash('Reporte en proceso, vuelva a revisar en unos minutos')
        return jsonify({'job': entry_id})

    company = get_company_info()
    header = [
        f"Empresa: {company.get('name', '')}",
        f"Rango: {(fecha_inicio or 'Todas')} - {(fecha_fin or 'Todas')}",
        f"Generado: {dom_now().strftime('%Y-%m-%d %H:%M')} por {user}",
    ]
    current_app.logger.info(
        "export user=%s company=%s formato=%s tipo=%s filtros=%s",
        user,
        current_company_id(),
        formato,
        tipo,
        filtros,
    )

    if formato == 'csv':
        if current_app.testing:
            output = StringIO()
            writer = csv.writer(output)
            for h in header:
                writer.writerow([h])
            if tipo == 'resumen':
                writer.writerow(['Categoría', 'Cantidad', 'Total'])
                summary = (
                    company_query(InvoiceItem)
                    .join(Invoice)
                    .with_entities(
                        InvoiceItem.category,
                        func.count(InvoiceItem.id),
                        func.sum(InvoiceItem.unit_price * InvoiceItem.quantity - InvoiceItem.discount),
                    )
                    .group_by(InvoiceItem.category)
                )
                if start:
                    summary = summary.filter(Invoice.date >= start)
                if end:
                    summary = summary.filter(Invoice.date <= end)
                if estado:
                    summary = summary.filter(Invoice.status == estado)
                for cat, cnt, tot in summary:
                    writer.writerow([cat or 'Sin categoría', cnt, f"{tot or 0:.2f}"])
            else:
                writer.writerow(['Cliente', 'Fecha', 'Estado', 'Total'])
                for inv in q.options(
                    joinedload(Invoice.client),
                    load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
                ):
                    writer.writerow([
                        inv.client.name if inv.client else '',
                        inv.date.strftime('%Y-%m-%d'),
                        inv.status or '',
                        f"{inv.total:.2f}",
                    ])
            mem = BytesIO()
            mem.write(output.getvalue().encode('utf-8'))
            mem.seek(0)
            log_export(user, formato, tipo, filtros, 'success')
            return send_file(mem, mimetype='text/csv', as_attachment=True, download_name='reportes.csv')
        app_obj = current_app._get_current_object()
        def generate_csv():
            with app_obj.app_context():
                sio = StringIO()
                writer = csv.writer(sio)
                for h in header:
                    writer.writerow([h])
                if tipo == 'resumen':
                    writer.writerow(['Categoría', 'Cantidad', 'Total'])
                    yield sio.getvalue(); sio.seek(0); sio.truncate(0)
                    summary = (
                        company_query(InvoiceItem)
                        .join(Invoice)
                        .with_entities(
                            InvoiceItem.category,
                            func.count(InvoiceItem.id),
                            func.sum(InvoiceItem.unit_price * InvoiceItem.quantity - InvoiceItem.discount),
                        )
                        .group_by(InvoiceItem.category)
                    )
                    if start:
                        summary = summary.filter(Invoice.date >= start)
                    if end:
                        summary = summary.filter(Invoice.date <= end)
                    if estado:
                        summary = summary.filter(Invoice.status == estado)
                    for cat, cnt, tot in summary:
                        writer.writerow([cat or 'Sin categoría', cnt, f"{tot or 0:.2f}"])
                        yield sio.getvalue(); sio.seek(0); sio.truncate(0)
                else:
                    writer.writerow(['Cliente', 'Fecha', 'Estado', 'Total'])
                    yield sio.getvalue(); sio.seek(0); sio.truncate(0)
                    stream_q = q.options(
                        joinedload(Invoice.client),
                        load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
                    ).yield_per(100)
                    for inv in stream_q:
                        writer.writerow([
                            inv.client.name if inv.client else '',
                            inv.date.strftime('%Y-%m-%d'),
                            inv.status or '',
                            f"{inv.total:.2f}",
                        ])
                        yield sio.getvalue(); sio.seek(0); sio.truncate(0)

        log_export(user, formato, tipo, filtros, 'success')
        headers = {
            'Content-Disposition': 'attachment; filename=reportes.csv'
        }
        return Response(generate_csv(), mimetype='text/csv', headers=headers)

    invoices = q.options(
        joinedload(Invoice.client),
        load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
    ).all()

    if formato == 'xlsx':
        if Workbook is None:
            mem = BytesIO()
            mem.write(b'')
            mem.seek(0)
            return send_file(mem, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='reportes.xlsx')
        wb = Workbook()
        ws = wb.active
        row = 1
        for h in header:
            ws.cell(row=row, column=1, value=h)
            row += 1
        if tipo == 'resumen':
            ws.append(['Categoría', 'Cantidad', 'Total'])
            summary = (
                company_query(InvoiceItem)
                .join(Invoice)
                .with_entities(InvoiceItem.category, func.count(InvoiceItem.id), func.sum(InvoiceItem.unit_price * InvoiceItem.quantity - InvoiceItem.discount))
                .group_by(InvoiceItem.category)
            )
            if start:
                summary = summary.filter(Invoice.date >= start)
            if end:
                summary = summary.filter(Invoice.date <= end)
            if estado:
                summary = summary.filter(Invoice.status == estado)
            for cat, cnt, tot in summary:
                ws.append([cat or 'Sin categoría', cnt, float(tot or 0)])
        else:
            ws.append(['Cliente', 'Fecha', 'Estado', 'Total'])
            for inv in invoices:
                ws.append([
                    inv.client.name if inv.client else '',
                    inv.date.strftime('%Y-%m-%d'),
                    inv.status or '',
                    float(inv.total),
                ])
        mem = BytesIO()
        wb.save(mem)
        mem.seek(0)
        log_export(user, formato, tipo, filtros, 'success')
        return send_file(
            mem,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='reportes.xlsx',
        )

    if formato == 'pdf':
        items = [
            {
                'code': inv.id,
                'reference': inv.client.name if inv.client else '',
                'product_name': inv.date.strftime('%d/%m/%Y'),
                'unit': inv.status or '',
                'unit_price': inv.total,
                'quantity': 1,
                'discount': 0,
            }
            for inv in invoices
        ]
        subtotal = sum(inv.total for inv in invoices)
        note = (
            f"Rango: {(fecha_inicio or 'Todas')} - {(fecha_fin or 'Todas')} | "
            f"Estado: {(estado or 'Todos')} | "
            f"Categoría: {(categoria or 'Todas')} | "
            f"Usuario: {user} | Facturas: {len(invoices)}"
        )
        pdf_data = generate_pdf_bytes(
            'Reporte de Facturas',
            company,
            {'name': '', 'address': '', 'phone': ''},
            items,
            subtotal,
            0,
            subtotal,
            note=note,
        )
        report_doc_number = datetime.now().strftime('%Y%m%d%H%M%S')
        archived_path = _archive_pdf_copy('reporte', report_doc_number, pdf_data, company_name=company.get('name'))
        log_export(user, formato, tipo, filtros, 'success', file_path=archived_path or 'memory:reportes.pdf')
        return _archive_and_send_pdf(
            doc_type='reporte',
            doc_number=report_doc_number,
            pdf_data=pdf_data,
            download_name='reportes.pdf',
            company_name=company.get('name'),
            archive=False,
        )

    return redirect(url_for('reportes'))


@app.route('/reportes/inventario/export')
def export_inventory():
    role = session.get('role')
    if role not in ('admin', 'manager', 'contabilidad'):
        return '', 403
    company_id = current_company_id()
    rows = (
        db.session.query(
            Product.code,
            Product.name,
            Warehouse.name,
            ProductStock.stock,
            ProductStock.min_stock,
        )
        .join(ProductStock, Product.id == ProductStock.product_id)
        .join(Warehouse, ProductStock.warehouse_id == Warehouse.id)
        .filter(ProductStock.company_id == company_id)
        .order_by(Product.name)
        .all()
    )
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Código', 'Producto', 'Almacén', 'Stock', 'Mínimo'])
    for code, name, wh, stock, min_stock in rows:
        writer.writerow([code or '', name or '', wh or '', stock, min_stock])
    mem = BytesIO()
    mem.write(output.getvalue().encode('utf-8'))
    mem.seek(0)
    return send_file(mem, mimetype='text/csv', as_attachment=True, download_name='inventario.csv')


@app.route('/reportes/exportes')
def export_history():
    q = company_query(ExportLog).order_by(ExportLog.created_at.desc())
    usuario = request.args.get('usuario')
    formato = request.args.get('formato')
    if usuario:
        q = q.filter(ExportLog.user == usuario)
    if formato:
        q = q.filter(ExportLog.formato == formato)
    logs = q.limit(100).all()
    return render_template('export_history.html', logs=logs)


@app.route('/docs')
def docs():
    return render_template('docs.html')


@app.route('/contabilidad')
def contabilidad():
    return render_template('contabilidad.html')


@app.route('/contabilidad/catalogo')
def contab_catalogo():
    return render_template('contabilidad_catalogo.html')


@app.route('/contabilidad/entradas')
def contab_entradas():
    return render_template('contabilidad_entradas.html')


@app.route('/contabilidad/estados')
def contab_estados():
    return render_template('contabilidad_estados.html')


@app.route('/contabilidad/libro-mayor')
def contab_libro_mayor():
    return render_template('contabilidad_libro_mayor.html')


@app.route('/contabilidad/impuestos')
def contab_impuestos():
    return render_template('contabilidad_impuestos.html')


@app.route('/contabilidad/balanza')
def contab_balanza():
    return render_template('contabilidad_balanza.html')


@app.route('/contabilidad/asignacion')
def contab_asignacion():
    return render_template('contabilidad_asignacion.html')


@app.route('/contabilidad/centro-costo')
def contab_centro_costo():
    return render_template('contabilidad_centro_costo.html')


@app.route('/contabilidad/reportes')
def contab_reportes():
    return render_template('contabilidad_reportes.html')


@app.route('/contabilidad/dgii')
def contab_dgii():
    return render_template('contabilidad_dgii.html')


@app.route('/api/recommendations')
def api_recommendations():
    """Return top product recommendations based on past orders."""
    return jsonify({'products': recommend_products(current_company_id())})

if __name__ == '__main__':
    debug = os.environ.get('FLASK_DEBUG', '').strip().lower() in {'1', 'true', 'yes', 'on'}
    host = os.environ.get('HOST', '127.0.0.1')
    port = int(os.environ.get('PORT', 5000))
    app.run(host=host, port=port, debug=debug)
